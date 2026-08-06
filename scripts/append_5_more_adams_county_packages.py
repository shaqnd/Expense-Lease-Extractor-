# -*- coding: utf-8 -*-
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from copy import copy
P="/home/user/Expense-Lease-Extractor-/Master_Property_Database.xlsx"
wb=load_workbook(P); NS="(not stated in document)"
MON='"$"#,##0.00'; MON0='"$"#,##0'; PSF='"$"#,##0.00'

def style_from(ws,r,nc): return [copy(ws.cell(row=r,column=c)._style) for c in range(1,nc+1)]
def put(ws,row,vals,styles,nf=None):
    for j,v in enumerate(vals,1):
        c=ws.cell(row=row,column=j); c.value=v
        if j-1<len(styles): c._style=copy(styles[j-1])
        if nf and j in nf: c.number_format=nf[j]
        c.alignment=Alignment(vertical="top",wrap_text=True)
def snapshot(ws,frm,nc):
    return [[(ws.cell(row=r,column=c).value,copy(ws.cell(row=r,column=c)._style),ws.cell(row=r,column=c).number_format)
             for c in range(1,nc+1)] for r in range(frm,ws.max_row+1)]
def restore(ws,at,buf):
    for i,rd in enumerate(buf):
        for j,(v,s,f) in enumerate(rd,1):
            c=ws.cell(row=at+i,column=j); c.value=v; c._style=copy(s); c.number_format=f
def insert_block(ws,at,rows,src,nc,nf=None):
    n=len(rows)
    moved=[m for m in list(ws.merged_cells.ranges) if m.min_row>=at]
    for m in moved: ws.unmerge_cells(str(m))
    buf=snapshot(ws,at,nc); st=style_from(ws,src,nc)
    for i,r in enumerate(rows): put(ws,at+i,r,st,nf)
    restore(ws,at+n,buf)
    for m in moved:
        ws.merge_cells(start_row=m.min_row+n,start_column=m.min_col,end_row=m.max_row+n,end_column=m.max_col)

# ---------------- SOURCE DOCUMENTS (insert before blank+notes)
ws=wb["Source Documents"]
srcrows=[
 [15,"R0103518.pdf","Tax appeal - income approach + 2 rent rolls + CY2023 income statement + 3 lease comps + LOA/40-parcel schedule","5977-5995 N. Broadway, Denver (First Industrial)","Assess. Yr 2024","Ryan LLC (Kim Lust) / CoStar"],
 [16,"R0103519.pdf","Valuation protest letter + LOA + 2023 & 2024 P&L + 5-date rent roll","5945 & 5957 Broadway (P & J Trujillo LLLP)","Tax Yr 2025 (val 6/30/2024)","1st Net Real Estate Services (Mike Walter)"],
 [17,"R0103704.pdf","County Commercial-Industrial Income & Expense Survey (handwritten)","5720 Washington St, Denver (CB Resources LLC)","2023-2024","Adams County Assessor / Owner"],
 [18,"R0103779.pdf","County Commercial-Industrial Income & Expense Survey (handwritten)","1890 E 58th Ave, Denver (Federal Partners LLC)","2023-2024","Adams County Assessor / Owner"],
 [19,"R0104122.pdf","SB 11-119 submission - 06/24 rent roll only","550 W 53rd Pl (WSDB 550 LLC)","2024","Stevens & Associates (Todd J. Stevens)"],
]
insert_block(ws,20,srcrows,6,6)

# ---------------- PROPERTY MASTER
ws=wb["Property Master"]; at=ws.max_row+1; st=style_from(ws,5,25)
pm=[
 ["R0103518","R0103518","R0103518","11459 5977-5995 N Broadway","5977-5995 N. Broadway","Denver","CO","80216 (per rent roll; cover sheet prints 80012)","Adams",
  "First Industrial Colorado","First Industrial LP","Industrial - Distribution","Multi (2 tenants)",4.960,216058,None,54431,1970,"3.97 : 1",
  "100% (rent rolls 1/1/2023 & 1/1/2024)",6885558,"2024",5794600,"First Industrial Realty Trust, Inc.","R0103518.pdf"],
 ["R0103519","R0103519","R0103519","5945 & 5957 Broadway","5945 Broadway","Denver","CO","80216 (est.)","Adams",
  "P & J Trujillo LLLP","P & J Trujillo LLLP","Commercial / Industrial (multi-tenant)","Multi (5 spaces)",None,None,None,65227,None,None,
  "100% - 'No vacant space' at all 5 rent-roll dates",7902868,"2025",5900000,"P & J Trujillo LLLP (Peggy L. Trujillo, GP)","R0103519.pdf"],
 ["R0103704","R0103704","0182511300068","5720 Washington St","5720 Washington St","Denver","CO","80216-1322","Adams",
  "CB Resources LLC","CB Resources LLC","Office / Warehouse ('office spaces')","Owner-occupied",None,None,None,18000,None,None,
  "100% owner-occupied",None,"2024 survey (2025 reappraisal)",None,"Owner","R0103704.pdf"],
 ["R0103779","R0103779","0182511400063","1890 E 58th Ave","1890 E 58th Ave","Denver","CO","80216","Adams",
  "Federal Partners LLC","Federal Partners LLC","Industrial - Warehouse","Part owner-occupied (~50%) / multi-tenant",None,None,None,None,None,None,
  "Owner occupies ~50% (47,000 SF); balance multi-tenant; 100% of building rented or subject to lease",None,"2024 survey (2025 reappraisal)",None,"Owner","R0103779.pdf"],
 ["R0104122","R0104122","0182515202002","550 W 53rd Pl","550 W 53rd Pl",NS,"CO",NS,"Adams",
  "WSDB 550 LLC","WSDB 550 LLC","Industrial / Retail (multi-tenant)","Multi (2 tenants)",None,None,None,127000,None,None,
  "100% (2 tenants on 06/24 rent roll)",None,"2025",None,"Owner","R0104122.pdf"],
]
for i,r in enumerate(pm): put(ws,at+i,r,st,{14:'0.000',15:'#,##0',16:'#,##0',17:'#,##0',18:'0',21:MON0,23:MON0})

# ---------------- TENANTS & LEASES
ws=wb["Tenants & Leases"]; at=ws.max_row+1; st=style_from(ws,6,13)
tl=[
 ["5977-5995 N Broadway (R0103518)","5977-81","EDGEBANDING SERVICES INC","Actual - industrial net",22904.43,274853.16,6.85,"11/1/2019","10/31/2024","Lease 8580-4 ver; orig. exec. 9/14/2005",None,22053.08,
  "40,140 SF. Rent roll as of 1/1/2023. Monthly net rent = base rent. O/E estimate $12,917.99/mo ($3.86 PSF). Move-in 11/1/2019."],
 ["5977-5995 N Broadway (R0103518)","005995","SPEC 7 INSULATION CO.","Actual - industrial net",7110.29,85323.48,8.41,"11/1/2021","10/31/2024","Lease 15546-2 ver; orig. exec. 6/16/2016",None,8426.58,
  "10,140 SF. Rent roll as of 1/1/2023. O/E estimate $3,320.93/mo ($3.93 PSF). Move-in 10/15/2016."],
 ["5977-5995 N Broadway (R0103518)","TOTAL 1/1/2023","2 tenants - 100% occupied","Actual - industrial net",30014.72,360176.64,7.16,"-","-","-",None,30479.66,
  "50,280 SF occupied, 0 SF vacant. Printed rent-roll totals as of 1/1/2023. NOTE: project SF on the rent roll is 50,280 while the appeal's stated NLA is 54,431 SF."],
 ["5977-5995 N Broadway (R0103518)","5977-81","EDGEBANDING SERVICES INC","Actual - industrial net",23591.45,283097.40,7.05,"11/1/2019","10/31/2024","-",None,22053.08,
  "40,140 SF. Rent roll as of 1/1/2024 (escalated). O/E estimate $15,263.38/mo ($4.56 PSF)."],
 ["5977-5995 N Broadway (R0103518)","005995","SPEC 7 INSULATION CO.","Actual - industrial net",7359.16,88309.92,8.71,"11/1/2021","10/31/2024","-",None,8426.58,
  "10,140 SF. Rent roll as of 1/1/2024. O/E estimate $3,999.94/mo ($4.73 PSF). Flagged '*' - a future lease version had been entered."],
 ["5977-5995 N Broadway (R0103518)","TOTAL 1/1/2024","2 tenants - 100% occupied","Actual - industrial net",30950.61,371407.32,7.38,"-","-","-",None,30479.66,
  "50,280 SF occupied, 0 SF vacant. Printed rent-roll totals as of 1/1/2024."],
 ["5945 & 5957 Broadway (R0103519)","Space 1 (19,712 SF)","(tenant not named on rent roll)","Actual - base + CAM",11909.34,142912.08,7.25,"-","-","-",4632.32,None,
  "Rent roll gives SF, monthly base rent and CAM only - no tenant names, lease dates or terms. Figures shown are the latest snapshot, 1/1/2025. Prior snapshots: 1/1/23 $8,845.31; 6/30/23 $9,856.00; 1/1/24 and 6/30/24 $11,088.00."],
 ["5945 & 5957 Broadway (R0103519)","Space 2 (12,000 SF)","(tenant not named on rent roll)","Actual - base + CAM",6268.78,75225.36,6.27,"-","-","-",2820.00,None,
  "1/1/2025 snapshot. Prior: 1/1/23 $6,008.92; 6/30/23 and 1/1/24 $6,186.19; 6/30/24 $6,368.78 (the 1/1/25 figure is $100 lower than 6/30/24, as printed)."],
 ["5945 & 5957 Broadway (R0103519)","Space 3 (10,000 SF)","(tenant not named on rent roll)","Actual - base + CAM",7918.33,95019.96,9.50,"-","-","-",2289.61,None,
  "1/1/2025 snapshot. Prior: 1/1/23 $7,463.33; 6/30/23 through 6/30/24 $7,687.50."],
 ["5945 & 5957 Broadway (R0103519)","Space 4 (11,015 SF)","(tenant not named on rent roll)","Actual - base + CAM",8772.04,105264.48,9.56,"-","-","-",2588.53,None,
  "1/1/2025 snapshot. Prior: 1/1/23 through 6/30/23 $8,220.86; 1/1/24 and 6/30/24 $8,467.78."],
 ["5945 & 5957 Broadway (R0103519)","Space 5 (12,500 SF)","(tenant not named on rent roll)","Actual - base + CAM",9897.92,118775.04,9.50,"-","-","-",2937.50,None,
  "1/1/2025 snapshot. Prior: 1/1/23 through 6/30/23 $9,329.17; 1/1/24 and 6/30/24 $9,609.38."],
 ["5945 & 5957 Broadway (R0103519)","TOTAL 1/1/2025","5 spaces - no vacant space","Actual - base + CAM",44766.41,537196.92,8.24,"-","-","-",15267.96,None,
  "65,227 SF across 5 spaces. Monthly base-rent totals by date: 1/1/23 $39,867.59; 6/30/23 $41,279.72; 1/1/24 $43,038.85; 6/30/24 $43,221.44; 1/1/25 $44,766.41. CAM total $15,267.96/mo at every date."],
 ["1890 E 58th Ave (R0103779)","Leased portion","(multi-tenant; tenants not named)","Actual - Gross",55906.00,670872.00,14.00,"-","-","Years; 9 years remaining as of 2024",None,None,
  "County I&E survey: scheduled rent $14.00/SF for 12 months beginning 1/1/2023; actual rent received $670,872 (= $55,906 x 12). Owner reports 100% of the building rented or subject to lease, while also reporting ~50% (47,000 SF) owner-occupancy - reproduced as printed. New lease(s) signed between 7/1/22 and 6/30/24."],
 ["550 W 53rd Pl (R0104122)","-","Monarch Metal Mfg.","Actual - base + CAM",32000.00,384000.00,4.41,"-","-","-",10000.00,None,
  "87,000 SF. 06/24 rent roll. Lease date left blank on the rent roll."],
 ["550 W 53rd Pl (R0104122)","-","Arc Thrift Store","Actual - base + CAM",20000.00,240000.00,6.00,"9/1/2020","9/30/2025","September 2020 - September 2025",5000.00,None,
  "40,000 SF. 06/24 rent roll. Lease dates shown only as month/year."],
 ["550 W 53rd Pl (R0104122)","TOTAL","2 tenants","Actual - base + CAM",52000.00,624000.00,4.91,"-","-","-",15000.00,None,
  "127,000 SF combined. Rent roll reports no vacancy, no lease PSF and no expiry for Monarch."],
]
for i,r in enumerate(tl): put(ws,at+i,r,st,{5:MON,6:MON,7:PSF,11:MON,12:MON})

# ---------------- ASSESSMENTS (block B then block A)
ws=wb["Assessments & Valuation"]
bB=[["5977-5995 N Broadway (R0103518)",489879,-39190,450689,-45069,405620,0.0700,5794600],
    ["5945 & 5957 Broadway (R0103519) - agent capitalization",None,None,None,None,550000,0.09331,5894300]]
insert_block(ws,32,bB,21,8,{2:MON0,3:MON0,4:MON0,5:MON0,6:MON0,7:'0.000%',8:MON0})
bA=[["5977-5995 N Broadway (R0103518)","2024",6885558,127.00,5794600,106.46,None,"R0103518"],
    ["5945 & 5957 Broadway (R0103519)","2025",7902868,121.16,5900000,90.45,None,"R0103519"]]
insert_block(ws,16,bA,6,8,{3:MON0,4:PSF,5:MON0,6:PSF,7:MON0})

# ---------------- I&E SUMMARY
ws=wb["Income-Expense Summary"]; at=19
ies=[
 ["5977-5995 N Broadway - CY2023",550031,250528,None,"Accrual. TRUE NOI - depreciation ($71,686) and lease-cost amortization ($9,843) sit below the NOI line; net income after those = $217,973. R0103518"],
 ["5945 & 5957 Broadway - CY2024",704810.59,524505.21,None,"Cash basis. NOT a true NOI: expenses include loan interest $140,115.06, income tax $71,185.00 and property tax $169,029.80. Agent's stated 2024 NOI excluding those = $560,634. R0103519"],
 ["5945 & 5957 Broadway - CY2023",655107.84,453235.50,None,"Cash basis. NOT a true NOI: includes loan interest $142,547.72, income tax $50,000.00, property tax $148,047.68. Agent's stated 2023 NOI = $542,466. R0103519"],
 ["5720 Washington St - 2023/2024 survey",None,71980.00,None,"100% owner-occupied, so no rental income is reported. Operating expenses only, and they were written into the TENANT column of the survey grid despite the owner-occupancy. R0103704"],
 ["1890 E 58th Ave - 12 mo from 1/1/2023",670872.00,244295.66,None,"County I&E survey, Gross lease. Owner-paid expenses only; janitorial and snow/trash were marked 'T' rather than given a dollar amount, and no property tax or management line was reported. R0103779"],
]
fixed=[]
for i,r in enumerate(ies):
    row=at+i; r=list(r); r[3]=f"=B{row}-C{row}" if (r[1] is not None and r[2] is not None) else "-"
    fixed.append(r)
insert_block(ws,at,fixed,5,5,{2:MON,3:MON,4:MON})

# ---------------- I&E DETAIL
ws=wb["Income-Expense Detail"]; D=[]
def blk(p,c,items,printed=None,label="TOTAL",note=""):
    D.append((p,c,items,printed,label,note))
blk("5977-5995 N Broadway - CY2023","Income",[
 ("401000 Base Rent Income",362048,""),("408010 S/L Rent Revenue Adjustment",-31913,"Straight-line adjustment"),
 ("410100 Monthly Estimated CAM",194867,""),("412060 Tenant Direct Billed CAM",2700,""),
 ("413030 Current Year CAM Accrual",22328,""),("415000 Prior Year CAM Reconciliation",0,"")],550031,"TOTAL REVENUE",
 "Printed total: $550,031. Sum of the printed line items is $550,030 - the source's own Total Rent Revenue subtotal ($330,136) is $1 above its components ($330,135). Both reproduced as printed.")
blk("5977-5995 N Broadway - CY2023","Expense",[
 ("500030 C/A Water/Sewer",10493,""),("500035 C/A Other Utilities",846,""),("500045 C/A Exterior Light",2860,""),
 ("500010 C/A Electric",0,""),("501035 C/A Fire Prevention",4891,""),("501040 C/A MC Landscaping",1737,""),
 ("501050 C/A MC HVAC",1500,""),("501060 C/A Other Exterior",5324,""),("501075 C/A Parking Lot",0,""),
 ("501090 C/A Roofing",1334,""),("501100 C/A Snow Removal",13465,""),("501135 C/A Other Repair/Maintenance",3268,""),
 ("501160 C/A Sweeping",873,""),("502025 C/A Real Estate Taxes",139832,""),
 ("502020 C/A Real Estate Tax Appeals",615,""),("580025 Real Estate Tax Appeal Contra",-615,"Offsets the appeal fee"),
 ("503010 C/A Insurance",7696,""),("504015 C/A Management Fees-Internal",22462,""),
 ("521060 D/B Other Exterior",2700,"Direct-billed"),("530010 L/L Utilities",100,""),
 ("530020 L/L Maintenance & Repairs",197,""),("530285 L/L Real Estate Tax True Up",9385,""),
 ("530290 L/L RE Tax Accrual",21566,"")],250528,"TOTAL OPERATING EXPENSES",
 "Printed total: $250,528. Sum of the printed line items is $250,529 - the source's own Total Common Area Repair/Maint. subtotal ($32,391) is $1 below its components ($32,392). Both reproduced as printed.")
blk("5945 & 5957 Broadway - CY2024","Income",[
 ("4000 Rent Income",521595.07,""),("4045 Cam Charge Income",183215.52,"")],704810.59,"TOTAL INCOME")
blk("5945 & 5957 Broadway - CY2024","Expense",[
 ("6500/6520 Property Tax",169029.80,""),("6220 Loan Interest",140115.06,"Debt service - excluded from a true NOI"),
 ("6510 Income Tax",71185.00,"Entity tax - excluded from a true NOI"),("6300 Repairs - Building Repairs",32738.00,""),
 ("6182 Insurance - Bus Owners Pkg",40823.02,""),("6235 Management Fees",30000.00,""),
 ("6333 Repairs - Painting & Drywall",19750.00,""),("6335 Repairs - Plbg, Htg & A/C",10771.62,""),
 ("6410 Water",3926.42,""),("6270 Professional Fees - Other",3608.29,""),("6280 Legal Fees",1500.00,""),
 ("6275 Accounting",800.00,""),("6232 Lawn & Garden",200.00,""),("6120 Bank Service Charges",33.00,""),
 ("6160 Dues and Subscriptions",25.00,"")],524505.21,"TOTAL EXPENSE")
blk("5945 & 5957 Broadway - CY2023","Income",[
 ("4000 Rent Income",495442.14,""),("4045 Cam Charge Income",159665.70,"")],655107.84,"TOTAL INCOME")
blk("5945 & 5957 Broadway - CY2023","Expense",[
 ("6520 Property Tax",148047.68,""),("6220 Loan Interest",142547.72,"Debt service - excluded from a true NOI"),
 ("6182 Insurance - Bus Owners Pkg",62455.69,""),("6510 Income Tax",50000.00,"Entity tax - excluded from a true NOI"),
 ("6235 Management Fees",33487.46,""),("6410 Water",8870.16,""),("6335 Repairs - Plbg, Htg & A/C",5086.62,""),
 ("6236 Maintenance - Parking Lot",1550.00,""),("6275 Accounting",1100.00,""),
 ("6310 Building Repairs",89.87,""),("6120 Bank Service Charges",0.30,"")],453235.50,"TOTAL EXPENSE")
blk("5720 Washington St - 2023/2024 survey","Expense",[
 ("Building Insurance",45000.00,""),("Utilities",16380.00,""),("Janitorial Services",6600.00,""),
 ("Snow/Trash Removal",4000.00,"")],None,"TOTAL REPORTED OPERATING EXPENSES",
 "No printed total on the survey form - this is the sum of the four handwritten entries. All four were written in the Tenant-Expenses column even though the owner reported 100% owner-occupancy.")
blk("1890 E 58th Ave - 12 mo from 1/1/2023","Income",[
 ("Actual rent received (12 months from 1/1/2023)",670872.00,"= $55,906.00 monthly x 12; scheduled rent $14.00/SF")],670872.00,"TOTAL INCOME")
blk("1890 E 58th Ave - 12 mo from 1/1/2023","Expense",[
 ("Utilities",122995.66,""),("Building Repair & Maintenance",100803.00,""),("Building Insurance",18057.00,""),
 ("Ground Maintenance",2440.00,""),("Janitorial Services",0,"Marked 'T' on the form, not a dollar amount"),
 ("Snow/Trash Removal",0,"Marked 'T' on the form, not a dollar amount")],None,"TOTAL REPORTED OWNER EXPENSES",
 "No printed total on the survey form - sum of the owner/landlord column. Gross lease, so the owner pays expenses; no property-tax or management line was reported.")

at=ws.max_row+1; st=style_from(ws,5,5); CHECK=[]; row=at
for p,c,items,printed,label,note in D:
    s=row
    for n_,a,nt in items:
        put(ws,row,[p,c,n_,a,nt],st,{4:MON}); row+=1
    pn=(f"Printed total: ${printed:,.2f}" if printed is not None else "")
    put(ws,row,[p,c,label,f"=SUM(D{s}:D{row-1})",note or pn],st,{4:MON})
    ws.cell(row=row,column=3).font=Font(name="Arial",size=10,bold=True)
    ws.cell(row=row,column=4).font=Font(name="Arial",size=10,bold=True)
    CHECK.append((p,c,sum(a for _,a,_ in items),printed)); row+=1

# ---------------- RELATED PARCELS (First Industrial LOA schedule - Adams County only)
ws=wb["Related Parcels"]; at=ws.max_row+1; st=style_from(ws,5,5)
ws.cell(row=at,column=1,value="First Industrial Realty Trust - Adams County parcels on the R0103518 letter-of-authorization schedule (40 parcels total across 6 CO counties):").font=Font(name="Arial",size=10,bold=True,italic=True)
rp=[["First Industrial LP","R0103518","5977 Broadway St.","Adams County","CO"],
    ["First Industrial LP","R0103559","5952 Broadway St.","Adams County","CO"],
    ["FR MASSACHUSETTS 7 LLC","R0103523","5909 Broadway St.","Adams County","CO"],
    ["First Industrial LP","R0024470","451 E. 124th Ave.","Adams County","CO"],
    ["First Industrial LP","R0024474","Thornton CO","Adams County","CO"],
    ["FR Massachusettes 7 LLC","R0111259","18150 E. 32nd Pl., Aurora","Adams County","CO"],
    ["FRASER AURORA LLC","R0084260","3400 Fraser St","Adams County","CO"],
    ["FR Park Plaza LLC","R0177662","21301 E 33rd Dr","Adams County","CO"],
    ["FR AURORA COMMERCE CENTER PHASE I LLC","R0198530","22300 E 26th Ave","Adams County","CO"],
    ["FR Aurora Commerce Center Phase I LLC","R0203646","22000 E 26th Ave","Adams County","CO"],
    ["FR Aurora Commerce Center Phase I LLC","R0203647","22010 E 26th Ave","Adams County","CO"],
    ["FR Aurora Commerce Center Phase I LLC","R0205071","22600 E 26th Ave","Adams County","CO"],
    ["FR 21110 E 31st LLC","R0083949","21110 E 31st Circle","Adams County","CO"],
    ["First industrial LP","R0180952","3350 ODESSA WAY","Adams County","CO"],
    ["FR 8000 EAST 96 LLC","R0215909","8000 E 96th Ave","Adams County","CO"],
    ["FR 8000 EAST 96 LLC","R0215910","8000 E 96th Ave","Adams County","CO"]]
for i,r in enumerate(rp): put(ws,at+1+i,r,st)

wb.save(P); print("SAVED\n")
print("%-42s %-8s %14s %14s %s"%("BLOCK","CAT","PY SUM","PRINTED","MATCH"))
bad=0
for p,c,s,pr in CHECK:
    if pr is None: print("%-42s %-8s %14.2f %14s %s"%(p[:42],c,s,"(none)","n/a - no printed total")); continue
    d=abs(s-pr); ok = d<0.005
    tag = "OK" if ok else ("ANNOTATED $%.0f source rounding"%d if d<=1.001 else "*** MISMATCH ***")
    if not ok and d>1.001: bad+=1
    print("%-42s %-8s %14.2f %14.2f %s"%(p[:42],c,s,pr,tag))
print("\nUNEXPLAINED MISMATCHES:",bad)
