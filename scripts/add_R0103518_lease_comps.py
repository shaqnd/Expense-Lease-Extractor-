# -*- coding: utf-8 -*-
from openpyxl import load_workbook
from openpyxl.styles import Font,PatternFill,Alignment,Border,Side
from openpyxl.utils import get_column_letter
from datetime import date
from copy import copy
F="Lease_Comparables_All.xlsx"; wb=load_workbook(F)
N=Font(name="Arial",size=10); I=Font(name="Arial",size=9,italic=True,color="FF595959"); B=Font(name="Arial",size=10,bold=True)
th=Side(style="thin",color="FFBFBFBF"); BOX=Border(left=th,right=th,top=th,bottom=th)
M2='"$"#,##0.00'; NUM='#,##0'; DT='mm/dd/yyyy'; MON='"$"#,##0'; GOLD="FFFFF2CC"
ws=wb["Lease Comparables"]
# locate last data row (row before the trailing italic note)
last=None
for r in range(5,ws.max_row+1):
    if ws.cell(row=r,column=2).value not in (None,""): last=r
S="R0103518 - 5977-5995 N Broadway"
NEW=[
 [S,1,"500 W 53rd Pl - 1st Floor Sublease","500 W 53rd Pl","Denver","CO","80216","Northwest Denver",25208,date(2024,2,1),date(2024,3,1),date(2027,7,1),"3 Years",7.95,"Effective (asking = starting = effective)","NNN","Sublease","New Lease",None,None,"Streech Properties",1971,"R0103518 p.17"],
 [S,2,"80 E 62nd Ave - 1st Floor Direct","80 E 62nd Ave","Denver","CO","80216","Northwest Denver",25000,date(2023,6,1),date(2023,6,1),date(2025,6,1),"2 Years",8.20,"Starting","NNN","Direct","Renewal",None,"M&M Cut Flora, LLC","Ares Industrial Real Estate In...",1960,"R0103518 p.17"],
 [S,3,"6401 Broadway - 1st Floor Direct","6401 Broadway","Denver","CO","80221","Northwest Denver",1017,date(2022,7,1),date(2022,9,1),None,None,9.50,"Starting","NNN","Direct","New Lease",None,"i9 Sports","Sixty-Four O One Broadway, LLC",1984,"R0103518 p.17"],
]
src=[copy(ws.cell(row=last,column=c)._style) for c in range(1,24)]
# push the trailing note down
noterows=[]
for r in range(last+1,ws.max_row+1):
    v=ws.cell(row=r,column=1).value
    if v: noterows.append(v); ws.cell(row=r,column=1).value=None
at=last+1
for i,row in enumerate(NEW):
    for j,v in enumerate(row,1):
        c=ws.cell(row=at+i,column=j); c.value=v; c._style=copy(src[j-1]); c.border=BOX; c.font=N
        c.alignment=Alignment(vertical="top",wrap_text=(j in(3,4,20,21)))
        if j in(10,11,12) and isinstance(v,date): c.number_format=DT
        if j==9: c.number_format=NUM
        if j==14: c.number_format=M2
        if j==19: c.number_format=MON
        if j==22: c.number_format='0'
        c.fill=PatternFill("solid",fgColor="FFFFFFFF")
r=at+len(NEW)+1
for t in noterows:
    ws.cell(row=r,column=1,value=t).font=I; r+=1
ws.auto_filter.ref=f"A4:{get_column_letter(23)}{at+len(NEW)-1}"

# stats tab
ws=wb["Set Statistics"]
tgt=None
for rr in range(1,ws.max_row+1):
    if str(ws.cell(row=rr,column=1).value or "").startswith("R0111559 / R0111560"): tgt=rr
src=[copy(ws.cell(row=tgt,column=c)._style) for c in range(1,8)]
buf=[[(ws.cell(row=rr,column=c).value,copy(ws.cell(row=rr,column=c)._style),ws.cell(row=rr,column=c).number_format)
      for c in range(1,8)] for rr in range(tgt+1,ws.max_row+1)]
new=[["R0103518 (Ryan/CoStar) - Rent $/SF",3,7.95,8.55,8.20,9.50,"No summary page in this package; computed here from the 3 comps. Simple mean $8.55; SF-weighted $8.10 (the 1,017 SF comp at $9.50 barely moves a weighted average)."]]
for i,row in enumerate(new):
    for j,v in enumerate(row,1):
        c=ws.cell(row=tgt+1+i,column=j); c.value=v; c._style=copy(src[j-1])
        c.alignment=Alignment(vertical="top",wrap_text=(j in(1,7)))
        if j==2: c.number_format=NUM
        if j in(3,4,5,6): c.number_format=M2
for i,rd in enumerate(buf):
    for j,(v,s,f) in enumerate(rd,1):
        c=ws.cell(row=tgt+1+len(new)+i,column=j); c.value=v; c._style=copy(s); c.number_format=f

# notes tab
ws=wb["Notes & Data Gaps"]
last=None
for rr in range(5,ws.max_row+1):
    if ws.cell(row=rr,column=1).value: last=rr
add=[("Scope update - 10 packages","R0103518","A second batch of 5 packages was added. Of those, only R0103518 carries lease comparables (3 of them). R0103519, R0103704, R0103779 and R0104122 contain none."),
 ("DATA GAP - no summary stats","R0103518 p.17","Ryan's comp page carries no summary-statistics block, so the low/average/median/high on the Set Statistics tab were computed here rather than transcribed."),
 ("NEAR-NEIGHBOUR, NOT THE SUBJECT","R0103518 comp 1","Comp 1 is 500 W 53rd Pl. Another property in this same batch, R0104122, is 550 W 53rd Pl - adjacent addresses but distinct parcels and distinct owners."),
 ("DATA GAP - concessions","R0103518","Free rent, escalations, TI allowance and office area are blank for all 3 comps; expiry is blank for comp 3."),
]
src=[copy(ws.cell(row=last,column=c)._style) for c in range(1,4)]
for i,(a,b_,c_) in enumerate(add):
    flag=a.startswith("DATA GAP") or a.startswith("NEAR")
    for j,v in enumerate([a,b_,c_],1):
        c=ws.cell(row=last+1+i,column=j); c.value=v; c._style=copy(src[j-1])
        c.font=B if (flag and j==1) else N; c.border=BOX
        c.alignment=Alignment(vertical="top",wrap_text=True)
        if flag: c.fill=PatternFill("solid",fgColor=GOLD)
wb.save(F)
import statistics as st
v=[7.95,8.20,9.50]; sf=[25208,25000,1017]
print("R0103518 comps: n=%d low %.2f mean %.4f median %.2f high %.2f  SF-weighted %.4f"%(
 len(v),min(v),st.mean(v),st.median(v),max(v),sum(a*b for a,b in zip(v,sf))/sum(sf)))
w=load_workbook(F); print("sheets:",w.sheetnames,"| comp rows:",sum(1 for r in range(5,w['Lease Comparables'].max_row+1) if w['Lease Comparables'].cell(row=r,column=2).value not in (None,"")))
