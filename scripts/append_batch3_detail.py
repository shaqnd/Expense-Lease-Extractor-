# -*- coding: utf-8 -*-
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
from copy import copy
P="/home/user/Expense-Lease-Extractor-/Master_Property_Database.xlsx"
wb=load_workbook(P); MON='"$"#,##0.00'
def style_from(ws,r,nc): return [copy(ws.cell(row=r,column=c)._style) for c in range(1,nc+1)]
def put(ws,row,vals,st,nf=None):
    for j,v in enumerate(vals,1):
        c=ws.cell(row=row,column=j); c.value=v
        if j-1<len(st): c._style=copy(st[j-1])
        if nf and j in nf: c.number_format=nf[j]
        c.alignment=Alignment(vertical="top",wrap_text=True)

ws=wb["Income-Expense Detail"]; D=[]
def blk(p,c,items,printed=None,label="TOTAL",note=""): D.append((p,c,items,printed,label,note))

blk("5710 E. 56th Avenue - CY2023","Expense",[
 ("Real Estate Taxes",137214,"$2.74 PSF"),("CAM - Reimbursable",76169,"$1.52 PSF"),
 ("CAM - Non-Reimbursable",1012,"$0.02 PSF"),("Reserves for Replacement",0,"Reported as $0")],214395,"TOTAL EXPENSES")
blk("5710 E. 56th Avenue - CY2024","Expense",[
 ("Real Estate Taxes",141168,"$2.82 PSF"),("CAM - Reimbursable",86637,"$1.73 PSF"),
 ("CAM - Non-Reimbursable",525,"$0.01 PSF"),("Reserves for Replacement",0,"Reported as $0")],228330,"TOTAL EXPENSES")
blk("Confluent Center 70 - CY2024","Income",[
 ("Rent",925921,"60.3% of EGI, $8.80 PSF"),("Reimbursements",609219,"39.7% of EGI, $5.79 PSF"),
 ("Other Income",908,"0.1% of EGI"),("Above/Below Market Rent Income",0,""),("Concessions",0,"")],1536047,"EFFECTIVE GROSS INCOME",
 "Printed total: $1,536,047. Sum of the printed components is $1,536,048 - a $1 rounding difference in the source. Both reproduced as printed.")
blk("Confluent Center 70 - CY2024","Expense",[
 ("Property Taxes",391900,"25.5% of EGI, $3.73 PSF"),("Reimbursable Operating Expenses",264713,"17.2% of EGI, $2.52 PSF"),
 ("Non-Recoverable Operating Expenses",125552,"8.2% of EGI, $1.19 PSF")],782165,"TOTAL OPERATING EXPENSES")
blk("Majestic Commercenter Bldg 5 - CY2024","Income",[
 ("4015 Industrial Rental Income",1140134.64,""),("4145 Real Estate Tax Reimbursement",349367.26,""),
 ("4150 Utilities Reimbursement",95769.03,""),("4135 Common Area Reimbursement",80939.49,""),
 ("4175 Reserve Income",71340.84,""),("4160 Management Fee Reimbursement",53004.09,""),
 ("4035 Rent Concessions",39276.22,"Positive as printed"),("4140 Insurance Reimbursement",26275.98,""),
 ("4158 Direct Tenant Reimbursement",1651.74,""),("4156 Tenant TI Reimb Income",-21.94,"Credit"),
 ("4040 Rent Amortization",-57376.98,"Credit")],1800360.37,"TOTAL REVENUE")
blk("Majestic Commercenter Bldg 5 - CY2024","Expense",[
 ("5500 Real Estate Taxes",349322.30,""),("5400 Management Fees",52858.61,""),("5637 Utilities-DTR",43479.56,""),
 ("5320 Parking Lot",30790.02,""),("5635 Utilities-Water",22373.45,""),("5115 Insurance-Fire & Extended",20905.00,""),
 ("5370 Snow Removal",19643.13,""),("5290 Landscaping",17060.13,""),("5620 Utilities-Gas/Heating",16533.91,""),
 ("5610 Utilities-Electricity",9700.66,""),("5350 Roof",8159.28,""),("5250 Fire Life Safety",7632.19,""),
 ("8445 Prof Service-Legal (Leasing)",6720.00,""),("5275 HVAC",5395.32,""),("5120 Insurance-Liability",5268.00,""),
 ("5615 Utilities-Fire Service",2246.08,""),("5312 Other-DTR",2679.52,""),("5380 Sweeping",2155.91,""),
 ("5630 Utilities-Telephone",1752.90,""),("5385 Window Washing",1366.00,""),("5305 Lighting",921.76,""),
 ("8450 Prof Service-Other Environ",447.63,""),("5335 Plumbing",288.85,""),("8435 Prof Service-Other",225.00,""),
 ("5205 Backflow Testing & Repair",180.00,""),("5285 Keys/Card Access",179.58,""),("5280 Janitorial Services",163.04,""),
 ("5365 Signs",160.95,""),("5235 Door Repairs",130.00,""),("8420 Prof Service-Legal",80.68,"")],628819.46,"TOTAL OPERATING EXPENSES")
blk("3420 Lisbon Street - CY2024","Income",[
 ("4015 Industrial Rental Income",280446.36,"Ties exactly to the 06/01/24 rent roll annual rent"),
 ("4145 Real Estate Tax Reimbursement",149606.51,""),("4135 Common Area Reimbursement",29118.48,""),
 ("4160 Management Fee Reimbursement",14616.96,""),("4175 Reserve Income",13579.56,""),
 ("4140 Insurance Reimbursement",9974.00,""),("4158 Direct Tenant Reimbursement",261.67,""),
 ("4040 Rent Amortization",-13121.16,"Credit")],484482.38,"TOTAL REVENUE")
blk("3420 Lisbon Street - CY2024","Expense",[
 ("5500 Real Estate Taxes",150063.24,""),("5400 Management Fees",14140.06,""),("5290 Landscaping",13736.12,""),
 ("5115 Insurance-Fire & Extended",8951.00,""),("5320 Parking Lot",5065.03,""),("5250 Fire Life Safety",3491.10,""),
 ("5370 Snow Removal",2923.75,""),("5350 Roof",1794.77,""),("5120 Insurance-Liability",1189.00,""),
 ("5620 Utilities-Gas/Heating",1081.59,""),("5280 Janitorial Services",750.00,""),("5380 Sweeping",491.25,""),
 ("8450 Prof Service-Other Environ",447.63,""),("5385 Window Washing",296.00,""),("5365 Signs",272.45,""),
 ("5205 Backflow Testing & Repair",270.00,""),("6290 Landscaping",316.50,""),("5312 Other-DTR",261.67,""),
 ("8225 Tenant Relations",126.90,""),("8420 Prof Service-Legal",80.68,""),("5305 Lighting",16.15,""),
 ("5285 Keys/Card Access",11.86,""),("5275 HVAC",0,"No amount printed")],205776.75,"TOTAL OPERATING EXPENSES")

at=ws.max_row+1; st=style_from(ws,5,5); CHECK=[]; row=at
for p,c,items,printed,label,note in D:
    s=row
    for n_,a,nt in items: put(ws,row,[p,c,n_,a,nt],st,{4:MON}); row+=1
    pn=(f"Printed total: ${printed:,.2f}" if printed is not None else "")
    put(ws,row,[p,c,label,f"=SUM(D{s}:D{row-1})",note or pn],st,{4:MON})
    ws.cell(row=row,column=3).font=Font(name="Arial",size=10,bold=True)
    ws.cell(row=row,column=4).font=Font(name="Arial",size=10,bold=True)
    CHECK.append((p,c,sum(a for _,a,_ in items),printed)); row+=1

# Related parcels
ws=wb["Related Parcels"]; at=ws.max_row+1; st=style_from(ws,5,5)
ws.cell(row=at,column=1,value="Majestic Commercenter / Majestic Realty - further Adams County buildings protested by Sterling (2025):").font=Font(name="Arial",size=10,bold=True,italic=True)
for i,r in enumerate([["Majestic Commercenter Phase 9","R0191099","3559 N Himalaya Road, Building 5, Aurora","Adams County","CO"],
                      ["Majestic Lisbon Buildings, LLC","R0200721","3420 N. Lisbon Street, Aurora","Adams County","CO"]]):
    put(ws,at+1+i,r,st)

wb.save(P); print("SAVED\n")
print("%-46s %-8s %14s %14s %s"%("BLOCK","CAT","PY SUM","PRINTED","MATCH")); bad=0
for p,c,s,pr in CHECK:
    if pr is None: print("%-46s %-8s %14.2f %14s n/a"%(p[:46],c,s,"(none)")); continue
    d=abs(s-pr); ok=d<0.005
    tag="OK" if ok else ("ANNOTATED $%.0f source rounding"%d if d<=1.001 else "*** MISMATCH ***")
    if not ok and d>1.001: bad+=1
    print("%-46s %-8s %14.2f %14.2f %s"%(p[:46],c,s,pr,tag))
print("\nUNEXPLAINED MISMATCHES:",bad)
