# -*- coding: utf-8 -*-
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from copy import copy

P="/home/user/Expense-Lease-Extractor-/Master_Property_Database.xlsx"
wb=load_workbook(P)
NS="(not stated in document)"

def style_from(ws,src_row,ncols):
    return [copy(ws.cell(row=src_row,column=c)._style) for c in range(1,ncols+1)]

def put(ws,row,vals,styles,numfmts=None):
    for j,v in enumerate(vals,1):
        c=ws.cell(row=row,column=j)
        c.value=v          # direct assign: cell(value=None) does NOT clear an existing value
        if j-1 < len(styles): c._style=copy(styles[j-1])
        if numfmts and (j in numfmts): c.number_format=numfmts[j]
        c.alignment=Alignment(vertical="top",wrap_text=True)

def snapshot(ws,frm,ncols):
    """Read values+styles+formats from `frm`..max_row so we can rewrite them lower down."""
    buf=[]
    for r in range(frm,ws.max_row+1):
        buf.append([(ws.cell(row=r,column=c).value,
                     copy(ws.cell(row=r,column=c)._style),
                     ws.cell(row=r,column=c).number_format) for c in range(1,ncols+1)])
    return buf

def restore(ws,at,buf):
    for i,rowdata in enumerate(buf):
        for j,(v,stl,nf) in enumerate(rowdata,1):
            c=ws.cell(row=at+i,column=j)
            c.value=v      # direct assign so None genuinely clears shifted-over content
            c._style=copy(stl); c.number_format=nf

def insert_block(ws,at,rows,src_row,ncols,numfmts=None):
    """Explicit shift-down. openpyxl's insert_rows silently drops merged rows, so
    merges at/below the insertion point are released, the block is rewritten, and
    the merges are re-applied at their shifted positions."""
    n=len(rows)
    moved=[mr for mr in list(ws.merged_cells.ranges) if mr.min_row>=at]
    for mr in moved: ws.unmerge_cells(str(mr))
    buf=snapshot(ws,at,ncols)
    st=style_from(ws,src_row,ncols)
    for i,r in enumerate(rows): put(ws,at+i,r,st,numfmts)
    restore(ws,at+n,buf)
    for mr in moved:
        ws.merge_cells(start_row=mr.min_row+n,start_column=mr.min_col,
                       end_row=mr.max_row+n,end_column=mr.max_col)

MON='"$"#,##0.00'; MON0='"$"#,##0'; PSF='"$"#,##0.00'

# ---------------------------------------------------------------- SOURCE DOCS
ws=wb["Source Documents"]
rows=[
 [10,"R0110351.pdf","Tax appeal - income approach + rent roll + 2yr income statements + LOA","18300 E 28th Ave, Aurora (ASB Real Estate)","Tax Yr 2025 (val 6/30/2024)","Ryan LLC (Ethan Horn) / CoStar"],
 [11,"R0111559.pdf","Assessor protest letter - income approach, base-period leases, rent roll, 2yr I&E","3590 N Himalaya Rd, Bldg 6, Aurora","Assess. Yr 2026 (val 6/30/2024)","Sterling Property Tax Specialists (Brenda L. Fearn)"],
 [12,"R0111560.pdf","Assessor protest letter - income approach, excess-vacancy discounting, rent roll, 2yr I&E","20320 E. 36th Dr, Bldg 7, Aurora","Assess. Yr 2026 (val 6/30/2024)","Sterling Property Tax Specialists (Brenda L. Fearn)"],
 [13,"R0114026.pdf","SB 11-119 submission - 2023 & 2024 P&L (excerpt) + 06/24 rent roll","6051 Washington St, Adams County","2023-2024","Stevens & Associates (Todd J. Stevens)"],
 [14,"R0121833.pdf","Real property summary analysis - income approach, rent roll, 2yr P&L, executed lease","5360 Washington St (Anchor Business Park), Denver","As of 1/1/2025 (2023-2024 P&L)","Joseph C. Sansone Company"],
]
insert_block(ws,15,rows,6,6)

# ---------------------------------------------------------------- PROPERTY MASTER
ws=wb["Property Master"]
pm=[
 ["R0110351","R0110351","R0110351","18300 E 28th Ave","18300 E 28th Ave","Aurora","CO","80011 (est.)","Adams",
  "ASB Real Estate","18300 E 28TH OWNER LLC","Industrial - Warehouse/Distribution","Single tenant",4.83,None,None,85604,2000,None,
  "100% (6/30/2024)",10683449,"2025",7939000,"Allegiance / ASB Allegiance Fund","R0110351.pdf"],
 ["R0111559","R0111559","R0111559","Majestic Commercenter - Building 6","3590 N Himalaya Road","Aurora","CO","80011","Adams",
  "Commercenter #6 LLC","Commercenter #6 LLC","Industrial - Warehouse","Multi (3 tenants)",None,None,None,125000,1998,None,
  "100% (6/30/2024 DOV)",None,"2026",8754500,"Majestic Realty Co. (Nicole Creighton, PM)","R0111559.pdf"],
 ["R0111560","R0111560","R0111560","Majestic Commercenter - Building 7","20320 E. 36th Drive","Aurora","CO","80011","Adams",
  "Commercenter #7 LLC","Commercenter #7 LLC","Industrial - Warehouse","Single tenant (vacant at DOV)",None,None,None,200000,1998,None,
  "0% - 100% vacant (6/30/2024 DOV)",None,"2026",10345200,"Majestic Realty Co. (Nicole Creighton, PM)","R0111560.pdf"],
 ["R0114026","R0114026","01825-10-1-02-023","6051 Washington St","6051 Washington St",NS,"CO",NS,"Adams",
  "Friesen-Washington Street LLC","Friesen-Washington Street LLC","Commercial / Industrial (multi-tenant)","Multi (3 tenants)",None,None,None,None,None,None,
  NS,None,"2025",None,"Owner (Shannon Friesen, bookkeeping)","R0114026.pdf"],
 ["R0121833","R0121833","25-31459-0001-CO (PTR); Appeal 159402","Anchor Business Park","5360 Washington Street","Denver","CO","80216","Adams",
  "Anchor Business Park LLC","Anchor Business Park LLC","Industrial - Warehouse","Multi (4 units / 5 leases)",None,None,36472,36472,2001,None,
  "100% (5 leases on 2024 rent roll)",5487982,"2025",3919000,"Bart S. Hansen (Property Manager)","R0121833.pdf"],
]
at=ws.max_row+1
st=style_from(ws,5,25)
for i,r in enumerate(pm):
    put(ws,at+i,r,st,{14:'0.00',15:'#,##0',16:'#,##0',17:'#,##0',18:'0',21:MON0,23:MON0})

# ---------------------------------------------------------------- TENANTS & LEASES
ws=wb["Tenants & Leases"]
tl=[
 ["18300 E 28th Ave (R0110351)","0100 (Bldg 01115B)","Old West Mattress Co., LLC","Actual - NNN (industrial)",45512.79,546153.48,6.38,"10/10/2018","4/30/2028","-",34564.79,"-",
  "85,604 SF, 100% occupied; rent roll as of 6/30/2024. $34,564.79/mo is cost recovery, not CAM. Scheduled base-rent increases: 2/1/2025 $46,654.18 ($6.54 PSF); 2/1/2026 $47,795.57 ($6.70); 2/1/2027 $49,008.29 ($6.87); 2/1/2028 $50,221.01 ($7.04). Recovery categories eff. 1/1/2025: CAM $1,839.61, INS $3,338.15, RET $23,941.67 per month."],
 ["Majestic Commercenter Bldg 6 (R0111559)","-","Exhibits USA, Inc.","Actual - Multi-Tenant Net",22933.42,275201.04,4.9943,"7/13/2007","7/31/2027","-","-","-","55,103 SF = 44.08% of building. Rent roll as of 06/30/22."],
 ["Majestic Commercenter Bldg 6 (R0111559)","-","Colorado Roofing Products, LLC","Actual - Multi-Tenant Net",20855.35,250264.20,5.7434,"3/1/2015","2/28/2025","-","-","-","43,574 SF = 34.86% of building. Rent roll as of 06/30/22."],
 ["Majestic Commercenter Bldg 6 (R0111559)","Suite 200","Invisible Structures, Inc.","Actual - Multi-Tenant Net",14428.48,173141.76,6.5776,"2/1/2020","1/31/2027","-","-","-","26,323 SF = 21.06% of building. Rent roll as of 06/30/22."],
 ["Majestic Commercenter Bldg 6 (R0111559)","TOTAL","3 tenants - 100% occupied","Actual - Multi-Tenant Net",58217.25,698607.00,5.5889,"-","-","-","-","-","125,000 SF total, 0 SF vacant. Printed rent-roll totals."],
 ["Majestic Commercenter Bldg 7 (R0111560)","Whole","FedEx Ground Package System, Inc.","Actual - Single Tenant Net",105833.33,1269999.96,6.35,"4/15/2012","4/30/2024","-","-","-",
  "200,000 SF. Rent roll as of 06/30/22. Lease expired 4/30/2024, which is why the property was 100% vacant at the 6/30/2024 date of value."],
 ["6051 Washington St (R0114026)","-","Colorado Signs","Actual - lease rate & term only",None,None,9.75,"7/1/2021","6/30/2025","-","-","-",
  "Rent roll (06/24) reports only tenant, rate PSF and term - no SF or monthly rent. Corresponding P&L income account is named 'Colorado Wrap'."],
 ["6051 Washington St (R0114026)","-","CPI / Thorp","Actual - lease rate & term only",None,None,11.65,"7/1/2023","6/30/2030","-","-","-",
  "Rent roll (06/24) reports only tenant, rate PSF and term."],
 ["6051 Washington St (R0114026)","-","Touch Stone Granite & Marble","Actual (details not legible)",None,None,None,"-","-","-","-","-",
  "Third rent-roll row is obliterated by a black scan artifact; tenant identified only from the P&L rent accounts. Rate and term NOT recoverable."],
 ["Anchor Business Park (R0121833)","A","Anchor Building Products LLC","Actual - NNN",1399.00,16788.00,None,"-","N/A","-","-","-","2024 rent roll; date of change 4/1/24. Term end shown as N/A (owner-affiliated tenant)."],
 ["Anchor Business Park (R0121833)","A","Consolidated Electrical Distributors","Actual - NNN",5660.00,67920.00,None,"-","9/30/2026","-","-","-","2024 rent roll; date of change 10/1/24."],
 ["Anchor Business Park (R0121833)","B","Integrative Environmental Systems, LLC","Actual - NNN",5767.29,69207.48,9.50,"12/1/2023","11/30/2026","36 mo + option 12/1/2027-11/30/2030","-",5767.29,
  "7,285 SF at $9.50/SF NNN with 3% annual escalations (executed lease dated 10/31/2023). Lease Table 1: Yr1 $9.50/$5,767.29/mo; Yr2 $9.79/$5,943.35; Yr3 $10.08/$6,119.40; total $213,960.48. First-month charges $2,872.60; late fee $250. Rent roll shows term end 10/31/2026 vs 11/30/2026 in the lease body, and Table 1 periods start 11/1/2023 vs the 12/1/2023 commencement - reproduced as printed (source contains typos in the Table 1 period column)."],
 ["Anchor Business Park (R0121833)","C","Juniper Overland","Actual - NNN",5640.00,67680.00,None,"-","3/31/2026","-","-","-","2024 rent roll; date of change 4/1/24."],
 ["Anchor Business Park (R0121833)","D","Restoration Logistics","Actual - NNN",7190.50,86286.00,None,"-","3/31/2025","-","-","-","2024 rent roll; date of change 4/1/24."],
 ["Anchor Business Park (R0121833)","TOTAL","5 leases - all NNN","Actual - NNN",25656.79,307881.48,8.44,"-","-","-","-","-","Printed rent-roll totals, signed Bart S. Hansen (Property Manager) 4/1/24. PSF = $307,881.48 / 36,472 SF."],
]
at=ws.max_row+1
st=style_from(ws,6,13)
for i,r in enumerate(tl):
    put(ws,at+i,r,st,{5:MON,6:MON,7:PSF,11:MON,12:MON})

# ---------------------------------------------------------------- ASSESSMENTS (B then A)
ws=wb["Assessments & Valuation"]
blockB=[
 ["18300 E 28th Ave (R0110351)",599228,-59923,539305,-43144,496161,0.0625,7939000],
 ["Majestic Commercenter Bldg 6 (R0111559)",781250,-78125,703125,-35156,667969,0.0763,8754500],
 ["Majestic Commercenter Bldg 7 (R0111560)",1250000,-125000,1125000,-56250,1068750,0.0763,14007200],
 ["Anchor Business Park - market (R0121833)",346484,-17324,329160,-49374,279786,0.0714,3919000],
 ["Anchor Business Park - actual 2024 (R0121833)",457717,None,457717,-110337,347380,0.0990,3510000],
 ["Anchor Business Park - actual 2023 (R0121833)",401204,None,401204,-119063,282140,0.0990,2851000],
]
insert_block(ws,21,blockB,16,8,{2:MON0,3:MON0,4:MON0,5:MON0,6:MON0,7:'0.00%',8:MON0})
blockA=[
 ["18300 E 28th Ave (R0110351)","2025",10683449,124.80,7939000,92.74,8100000,"R0110351"],
 ["Majestic Commercenter Bldg 6 (R0111559)","2026",None,None,8754500,70.04,None,"R0111559"],
 ["Majestic Commercenter Bldg 7 (R0111560)","2026",None,None,10345200,51.73,None,"R0111560"],
 ["6051 Washington St (R0114026)","2025",None,None,None,None,None,"R0114026"],
 ["Anchor Business Park (R0121833)","2025",5487982,150.00,3919000,107.45,None,"R0121833"],
]
insert_block(ws,11,blockA,6,8,{3:MON0,4:PSF,5:MON0,6:PSF,7:MON0})

# ---------------------------------------------------------------- I&E SUMMARY
ws=wb["Income-Expense Summary"]
at=9
ies=[
 ["18300 E 28th Ave - CY2024",943294.95,390362.08,None,"Accrual. True NOI: no debt service, depreciation or capex in operating expenses. R0110351"],
 ["18300 E 28th Ave - CY2023",883682.08,361874.61,None,"Accrual. True NOI. R0110351"],
 ["Majestic Commercenter Bldg 6 - CY2024",1631357.84,642378.16,None,"Operating income before interest ($276,132.81) and other/depreciation ($548,107.94); net income after those = $164,738.93. R0111559"],
 ["Majestic Commercenter Bldg 6 - CY2023",1809817.41,975170.45,None,"Operating income before interest ($284,797.60) and other/depreciation ($592,527.05); net income after those = ($42,677.69). R0111559"],
 ["Majestic Commercenter Bldg 7 - CY2024",601617.52,856598.01,None,"Operating LOSS - FedEx lease expired 4/30/2024. Before interest ($443,845.39) and other/depr ($721,647.35); net = ($1,420,473.23). R0111560"],
 ["Majestic Commercenter Bldg 7 - CY2023",2280295.00,1162402.67,None,"Operating income before interest ($457,772.92) and other/depr ($795,716.24); net = ($135,596.83). Expenses include $887,022.00 real estate taxes. R0111560"],
 ["6051 Washington St - CY2024",282582.82,125492.28,None,"Cash basis. NOT a true NOI: expenses include depreciation $78,017.21 + amortization $5,811.58 and lease commissions $34,701.00, while property tax is $0.00. R0114026"],
 ["6051 Washington St - CY2023",216726.00,None,None,"Cash basis, EXCERPT ONLY. Income total as printed; itemised rents total $214,254.27, so $2,471.73 sits on a page not included in the submission. Expense detail incomplete (only accounting $3,006.92, amortization $855.00, bank fees $30.00 shown). R0114026"],
 ["Anchor Business Park - CY2024",457717.43,290220.40,None,"Accrual. NOT a true NOI: includes interest expense $44,300.59. Total expense $289,750.40 + other expense $470.00. Net income $167,497.03. R0121833"],
 ["Anchor Business Park - CY2023",401203.65,285540.55,None,"Accrual. NOT a true NOI: includes interest expense $46,920.74. Net income $115,663.10. R0121833"],
]
fixed=[]
for i,r in enumerate(ies):
    row=at+i; r=list(r)
    r[3]= f"=B{row}-C{row}" if r[2] is not None else "-"
    fixed.append(r)
insert_block(ws,at,fixed,5,5,{2:MON,3:MON,4:MON})

# ---------------------------------------------------------------- I&E DETAIL
ws=wb["Income-Expense Detail"]
D=[]
def blk(prop,cat,items,printed=None,label="TOTAL"):
    D.append((prop,cat,items,printed,label))

# --- 18300 E 28th Ave CY2024
blk("18300 E 28th Ave - CY2024","Income",[
 ("Base Rent - Industrial",545012.10,""),("Recovery - Current Yr CAM/Operating Exp",21372.13,""),
 ("Recovery - Current Yr Insurance",37117.43,""),("Recovery - Current Yr Taxes",348999.96,""),
 ("Recovery - Prior Yr CAM/Operating Exp",7784.82,""),("Recovery - Prior Yr Insurance",197.31,""),
 ("Recovery - Prior Yr Taxes",-17188.80,"Credit")],943294.95,"TOTAL REVENUE")
blk("18300 E 28th Ave - CY2024","Expense",[
 ("Admin - General",8.21,""),("Admin - Full-Time Salaries",5638.74,""),("Admin - Office Supplies & Equipment",1.46,""),
 ("Professional Fee - MRI",3236.14,""),("Cleaning - Pest Control",165.00,""),
 ("Professional Fee - Sustainability",1627.76,""),("Bank Charges",2519.66,""),("Management Fee",8175.16,""),
 ("Real Estate Taxes",348999.96,""),("Prior Year Taxes",-17188.80,"Credit"),
 ("Insurance - Property",25578.36,""),("Insurance - Liability",3521.23,""),("Insurance - Other",8017.84,""),
 ("Nonrecoverable - R&M Other",61.36,"")],390362.08,"TOTAL OPERATING EXPENSES")
# --- 18300 E 28th Ave CY2023
blk("18300 E 28th Ave - CY2023","Income",[
 ("Base Rent - Industrial",531386.87,""),("Recovery - Current Yr CAM/Operating Exp",18538.35,""),
 ("Recovery - Current Yr Insurance",29594.51,""),("Recovery - Current Yr Taxes",308202.90,""),
 ("Recovery - Prior Yr CAM/Operating Exp",-766.53,"Credit"),("Recovery - Prior Yr Insurance",-312.24,"Credit"),
 ("Recovery - Prior Yr Taxes",-2961.78,"Credit")],883682.08,"TOTAL REVENUE")
blk("18300 E 28th Ave - CY2023","Expense",[
 ("Admin - General",57.10,""),("Admin - Full-Time Salaries",6546.36,""),("Admin - Office Supplies & Equipment",1.61,""),
 ("Professional Fee - MRI",2752.60,""),("Professional Fee - Audit and Tax Filing",7310.78,""),
 ("Professional Fee - Legal",518.50,""),("Professional Fee - Sustainability",915.86,""),("Bank Charges",768.07,""),
 ("Management Fee",7970.79,""),("Real Estate Taxes",308202.90,""),("Prior Year Taxes",-2961.78,"Credit"),
 ("Insurance - Property",20559.31,""),("Insurance - Liability",3820.70,""),("Insurance - Other",5411.81,"")],361874.61,"TOTAL OPERATING EXPENSES")
# --- Anchor Business Park CY2024
blk("Anchor Business Park - CY2024","Income",[
 ("Rental Income",295017.10,""),("Property Taxes - Tenant",118960.59,""),("Building Insurance",20106.00,""),
 ("Gas & Electric",8448.94,""),("Community Area Maintenance",5487.80,""),("Water & Sewer",5097.00,""),
 ("Trash",3600.00,""),("Misc. Tenant Repairs",1000.00,"")],457717.43,"TOTAL INCOME")
blk("Anchor Business Park - CY2024","Expense",[
 ("Property Tax Expense",133608.36,""),("Interest Expense - Loan",44300.59,"Debt service - excluded from a true NOI"),
 ("Property Management Fees",31500.00,""),("Building Maintenance",29130.84,""),("Insurance Expense",18520.09,""),
 ("Utilities",9783.05,""),("Trash Removal",5821.82,""),("Common Area Maintenance",5750.00,""),
 ("Water & Sewer Expense",4616.57,""),("Telephone Expense",1780.94,""),("Automobile Expense",1504.20,""),
 ("Accounting Expense",1062.50,""),("Adams Co Stormwater Utility Fee",886.00,""),("Fire Alarm Monitoring",830.44,""),
 ("Fire Alarm Inspection & Repairs",655.00,""),("Ask My Accountant (other expense)",470.00,"")],290220.40,"TOTAL EXPENSE (incl. other)")
# --- Anchor Business Park CY2023
blk("Anchor Business Park - CY2023","Income",[
 ("Rental Income",279618.69,""),("Property Taxes - Tenant",80179.09,""),("Building Insurance",18863.75,""),
 ("Gas & Electric",9403.07,""),("Community Area Maintenance",7474.61,""),("Water & Sewer",3485.67,""),
 ("Trash",2847.77,""),("Misc. Tenant Repairs",-669.00,"Credit")],401203.65,"TOTAL INCOME")
blk("Anchor Business Park - CY2023","Expense",[
 ("Property Tax Expense",117601.18,""),("Interest Expense - Loan",46920.74,"Debt service - excluded from a true NOI"),
 ("Building Maintenance",21081.48,""),("Property Management Fees",19500.00,""),("Utilities",18166.07,""),
 ("Commission Expense",17116.84,""),("Insurance Expense",15675.73,""),("Common Area Maintenance",6990.00,""),
 ("Tenant Repairs",5957.21,""),("Trash Removal",5314.77,""),("Water & Sewer Expense",3690.16,""),
 ("Automobile Expense",1955.46,""),("Telephone Expense",1629.46,""),("Accounting Expense",1525.00,""),
 ("Adams Co Stormwater Utility Fee",886.00,""),("Fire Alarm Inspection & Repairs",820.65,""),
 ("Fire Alarm Monitoring",673.92,""),("Office Supplies",25.88,""),("Bank Fees",10.00,"")],285540.55,"TOTAL EXPENSE")
# --- 6051 Washington CY2024
blk("6051 Washington St - CY2024","Income",[
 ("Rent - CPI",153780.00,""),("Rent - Colorado Wrap",63525.00,""),("Rent - Touch Stone",62480.00,""),
 ("Management Fee - CPI",1537.80,""),("Management Fee - Colorado Wrap",635.22,""),
 ("Management Fees - Touchstone",624.80,"")],282582.82,"TOTAL INCOME")
blk("6051 Washington St - CY2024","Expense",[
 ("Depreciation",78017.21,"Non-cash - excluded from a true NOI"),
 ("Lease Commissions",34701.00,"Capital/leasing cost - excluded from a true NOI"),
 ("Amortization expense",5811.58,"Non-cash - excluded from a true NOI"),("Accounting Expense",4675.00,""),
 ("Roof Maintenance - CAM",875.00,""),("Owner Expense",527.24,""),("Repair and Maintenance - CAM",525.25,""),
 ("Legal Expense",345.00,""),("Bank Fees",15.00,""),("Property Tax",0.00,"Reported as $0.00"),
 ("HVAC / Insurance / Landscape / Trash / Utilities / Water - CAM",0.00,"All reported as $0.00")],125492.28,"TOTAL EXPENSES")
# --- 6051 Washington CY2023 (excerpt)
blk("6051 Washington St - CY2023","Income",[
 ("Rent - CPI",124025.00,""),("Rent - Colorado Wrap",61425.84,""),("Rent - Touch Stone",28803.43,"")],None,
 "SUBTOTAL of itemised rents (printed 'Total for Income with sub-accounts' = $216,726.00; $2,471.73 is on a page not included)")
blk("6051 Washington St - CY2023","Expense",[
 ("Accounting Expense",3006.92,""),("Amortization expense",855.00,""),("Bank Fees",30.00,"")],None,
 "PARTIAL - expense detail truncated; no printed total in the submitted excerpt")
# --- Commercenter 6 / 7 (category level)
blk("Majestic Commercenter Bldg 6 - CY2024","Income",[
 ("Industrial Rental Income",816290.11,""),("Real Estate Tax Reimbursement",370458.18,""),
 ("Utilities Reimbursement",237036.40,""),("Common Area Reimbursement",112044.62,""),
 ("Management Fee Reimbursement",51268.75,""),("Reserve Income",34477.56,""),
 ("Insurance Reimbursement",18409.99,""),("Rent Above/Below Market",11134.68,""),
 ("Direct Tenant Reimbursement",3874.98,""),("Tenant TI Reimb Income",2545.32,""),
 ("Rent Amortization",-26182.75,"Credit")],1631357.84,"TOTAL REVENUE")
blk("Majestic Commercenter Bldg 6 - CY2024","Expense",[
 ("Real Estate Taxes",238177.79,""),("Utilities - DTR",190655.91,""),("Management Fees",48115.93,""),
 ("Parking Lot",27882.26,""),("Utilities - Water",26890.90,""),("Landscaping",19468.01,""),
 ("Insurance - Fire & Extended",14262.00,""),("Snow Removal",13056.61,""),("Utilities - Gas/Heating",12760.66,""),
 ("Fire Life Safety",8339.40,""),("Painting",6315.00,""),("Prof Service - Legal (Leasing)",4730.00,""),
 ("Roof",4541.24,""),("Prof Service - Tax Prep",4206.00,""),("Insurance - Liability",4138.00,""),
 ("Utilities - Electricity",3791.07,""),("Other-DTR",3661.72,""),("Janitorial Services",2100.00,""),
 ("Utilities - Telephone",1907.55,""),("Lighting",1731.19,""),("Sweeping",1694.89,""),
 ("Other line items (backflow 270.00, HVAC 616.79, keys 179.58, repairs 169.00, signs 600.45, window washing 460.00, other utilities 15.00, tenant relations 380.69, misc admin (0.01), legal 80.68, prof svc other 732.22, environmental 447.63)",3952.03,"Aggregated minor accounts")],642378.16,"TOTAL OPERATING EXPENSES")
blk("Majestic Commercenter Bldg 6 - CY2023","Income",[
 ("Industrial Rental Income",794984.93,""),("Real Estate Tax Reimbursement",556975.99,""),
 ("Utilities Reimbursement",224211.78,""),("Common Area Reimbursement",81142.15,""),
 ("Direct Tenant Reimbursement",48969.89,""),("Management Fee Reimbursement",39899.18,""),
 ("Reserve Income",34477.56,""),("Insurance Reimbursement",16596.00,""),
 ("Rent Above/Below Market",11134.68,""),("Rent Concessions",4125.91,""),
 ("Other Reimbursements",3757.50,""),("Tenant TI Reimb Income",2545.32,""),
 ("Rent Amortization",-9003.48,"Credit")],1809817.41,"TOTAL REVENUE")
blk("Majestic Commercenter Bldg 6 - CY2023","Expense",[
 ("Real Estate Taxes",556976.00,""),("Utilities - DTR",197089.44,""),("Management Fees",51072.56,""),
 ("Other-DTR",31648.03,""),("Utilities - Water",18658.55,""),("Utilities - Gas/Heating",18061.78,""),
 ("Parking Lot",15959.73,""),("Repairs",13081.96,""),("Insurance - Fire & Extended",12342.00,""),
 ("Snow Removal",12119.43,""),("Landscaping",10957.08,""),("Prof Service - Legal (Leasing)",6953.93,""),
 ("Fire Life Safety",5466.57,""),("Roof",5407.50,""),("Utilities - Electricity",4680.18,""),
 ("Insurance - Liability",4496.00,""),("Janitorial Services",2154.16,""),("Utilities - Telephone",1618.50,""),
 ("Sweeping",1473.97,""),("Bank Service Fees",1242.82,""),("Lighting",1023.14,""),
 ("Other line items (backflow, glass, painting, signs, window washing, other utilities, other prof svc, environmental)",2687.12,"Aggregated minor accounts")],975170.45,"TOTAL OPERATING EXPENSES")
blk("Majestic Commercenter Bldg 7 - CY2024","Income",[
 ("Industrial Rental Income",433916.64,""),("Other Reimbursements",158900.00,""),
 ("Common Area Reimbursement",30792.45,""),("Insurance Reimbursement",21542.79,""),
 ("Management Fee Reimbursement",12984.33,""),("Direct Tenant Reimbursement",12100.00,""),
 ("Reserve Income",8666.68,""),("Utilities Reimbursement",1113.72,""),
 ("Real Estate Tax Reimbursement",-53610.69,"Credit - reversal on tenant departure"),
 ("Rent Amortization",-24788.40,"Credit")],601617.52,"TOTAL REVENUE")
blk("Majestic Commercenter Bldg 7 - CY2024","Expense",[
 ("Property Taxes (6500)",416274.89,""),("Repairs",75423.11,""),("Fire Life Safety (6250)",59287.01,""),
 ("Concrete Repairs",32885.00,""),("Leasing - Ground Lease",29315.00,""),("Parking Lot (6320)",29326.63,""),
 ("Utilities - Water (6635)",21194.33,""),("Painting",20720.00,""),("Management Fees",20123.21,""),
 ("Other-DTR",19050.15,""),("Insurance - Fire & Extended",17999.00,""),("HVAC (6275)",17159.38,""),
 ("Utilities - Electricity",15456.85,""),("Snow Removal (5370)",13374.26,""),("Fire Life Safety (5250)",12824.29,""),
 ("Parking Lot (5320)",12938.36,""),("Janitorial Services",12258.24,""),("Utilities - Gas/Heating",10275.31,""),
 ("Roof (5350)",10428.00,""),("Landscaping (5290)",10027.04,""),("Snow Removal (6370)",8979.60,""),
 ("Insurance - Liability",6631.00,""),("Landscaping (6290)",6686.44,""),("HVAC (5275)",4732.00,""),
 ("Prof Service - Tax Prep",4206.00,""),("Utilities - Water (5635)",4131.02,""),
 ("Prof Service - Legal (Leasing)",2710.00,""),("Door Repairs",2575.31,""),("Plumbing",1944.32,""),
 ("Sweeping",2711.83,""),("Other minor accounts (lighting 484.80+852.67, signs 493.46+398.77, window washing 130.00+390.00, backflow 435.00, electrical 476.32, keys 779.00, pest 280.00, roof 6350 1,596.75, repairs 5345 797.50, other utilities 500.92, legal 80.68, prof svc other 732.22)",8428.09,"Aggregated"),
 ("Real Estate Taxes (5500)",-53477.66,"Credit - reversal")],856598.01,"TOTAL OPERATING EXPENSES")
blk("Majestic Commercenter Bldg 7 - CY2023","Income",[
 ("Industrial Rental Income",1291166.60,""),("Real Estate Tax Reimbursement",887022.00,""),
 ("Common Area Reimbursement",112314.39,""),("Management Fee Reimbursement",27573.73,""),
 ("Reserve Income",26000.04,""),("Rent Amortization",-63781.76,"Credit")],2280295.00,"TOTAL REVENUE")
blk("Majestic Commercenter Bldg 7 - CY2023","Expense",[
 ("Real Estate Taxes",887022.00,""),("Leasing - Ground Lease",87230.00,""),("Parking Lot",56960.31,""),
 ("Management Fees",39763.87,""),("Snow Removal",24526.92,""),("Insurance - Fire & Extended",15576.00,""),
 ("Landscaping",13090.67,""),("HVAC",9664.00,""),("Insurance - Liability",7206.00,""),
 ("Fire Life Safety",6193.09,""),("Roof",4195.00,""),("Repairs",3267.42,""),("Sweeping",2352.16,""),
 ("Lighting",1385.67,""),("Bank Service Fees",1274.17,""),("Prof Service - Other",733.55,""),
 ("Window Washing",520.00,""),("Backflow Testing & Repair",491.29,""),("Prof Service - Other Environ",400.00,""),
 ("Keys/Card Access",252.01,""),("Other-DTR",162.50,""),("Signs",81.71,""),("Painting",52.63,""),
 ("Utilities - Other / Misc admin",1.70,"")],1162402.67,"TOTAL OPERATING EXPENSES")

at=ws.max_row+1
st=style_from(ws,5,5)
CHECK=[]
row=at
for prop,cat,items,printed,label in D:
    start=row
    for name,amt,note in items:
        put(ws,row,[prop,cat,name,amt,note],st,{4:MON}); row+=1
    put(ws,row,[prop,cat,label,f"=SUM(D{start}:D{row-1})",
                (f"Printed total: ${printed:,.2f}" if printed is not None else "")],st,{4:MON})
    ws.cell(row=row,column=3).font=Font(name="Arial",size=10,bold=True)
    ws.cell(row=row,column=4).font=Font(name="Arial",size=10,bold=True)
    CHECK.append((prop,cat,sum(a for _,a,_ in items),printed))
    row+=1

# ---------------------------------------------------------------- RELATED PARCELS
ws=wb["Related Parcels"]
at=ws.max_row+1
st=style_from(ws,5,5)
rp=[["Commercenter #6 LLC","R0111559","3590 N Himalaya Road, Building 6, Aurora","Adams County","CO"],
    ["Commercenter #7 LLC","R0111560","20320 E. 36th Drive, Building 7, Aurora","Adams County","CO"]]
ws.cell(row=at,column=1,value="Majestic Commercenter - buildings protested together by Sterling Property Tax Specialists (2026):").font=Font(name="Arial",size=10,bold=True,italic=True)
for i,r in enumerate(rp): put(ws,at+1+i,r,st)

wb.save(P)
print("SAVED")
print()
print("%-46s %-8s %14s %14s %s"%("BLOCK","CAT","PY SUM","PRINTED","MATCH"))
bad=0
for prop,cat,s,p in CHECK:
    if p is None: print("%-46s %-8s %14.2f %14s %s"%(prop[:46],cat,s,"(none)","n/a - partial")); continue
    ok=abs(s-p)<0.005
    if not ok: bad+=1
    print("%-46s %-8s %14.2f %14.2f %s"%(prop[:46],cat,s,p,"OK" if ok else "*** MISMATCH ***"))
print()
print("MISMATCHES:",bad)
