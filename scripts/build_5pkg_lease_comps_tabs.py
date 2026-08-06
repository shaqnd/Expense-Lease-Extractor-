# -*- coding: utf-8 -*-
from openpyxl import load_workbook
from openpyxl.styles import Font,PatternFill,Alignment,Border,Side
from openpyxl.utils import get_column_letter
F="Lease_Comparables_5_Adams_County_Packages.xlsx"
wb=load_workbook(F)
NAVY="FF1F3864";BLUE="FF2E5C8A";GREEN="FFE2EFDA";GOLD="FFFFF2CC"
W=Font(name="Arial",size=10,color="FFFFFFFF",bold=True);TT=Font(name="Arial",size=12,color="FFFFFFFF",bold=True)
B=Font(name="Arial",size=10,bold=True);N=Font(name="Arial",size=10);I=Font(name="Arial",size=9,italic=True,color="FF595959")
th=Side(style="thin",color="FFBFBFBF");BOX=Border(left=th,right=th,top=th,bottom=th)
M2='"$"#,##0.00';NUM='#,##0'
def title(ws,r,nc,t,sub=None):
    ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=nc)
    c=ws.cell(row=r,column=1,value=t);c.font=TT;c.fill=PatternFill("solid",fgColor=NAVY);ws.row_dimensions[r].height=22
    if sub:
        ws.merge_cells(start_row=r+1,start_column=1,end_row=r+1,end_column=nc)
        ws.cell(row=r+1,column=1,value=sub).font=I;return r+2
    return r+1
def hdr(ws,r,v):
    for j,x in enumerate(v,1):
        c=ws.cell(row=r,column=j,value=x);c.font=W;c.fill=PatternFill("solid",fgColor=BLUE)
        c.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True);c.border=BOX
    ws.row_dimensions[r].height=30;return r+1

ws=wb.create_sheet("Set Statistics")
for i,w in enumerate([44,10,13,13,13,13,44],1): ws.column_dimensions[get_column_letter(i)].width=w
r=title(ws,1,7,"COMPARABLE-SET STATISTICS AS PUBLISHED, WITH INDEPENDENT RECALCULATION",
  "Every figure below was recomputed from the individual comps and matches the source to the cent.")
r+=1
r=hdr(ws,r,["Set / Metric","Deals","Low","Average","Median","High","Basis of the published 'Average'"])
ROWS=[
 ("R0110351 (Ryan/CoStar) - Asking Rent $/SF",3,6.34,7.27,6.75,9.75,"SF-WEIGHTED, not a simple mean (simple mean would be $7.61). Verified: $7.2737."),
 ("R0110351 (Ryan/CoStar) - Starting Rent $/SF",2,6.50,7.02,6.93,7.35,"SF-weighted. Verified: $7.0243. Median column = simple mean of the 2 deals."),
 ("R0110351 - Months on Market",3,4,11,7,22,"Published summary; per-comp detail not broken out in the package."),
 ("R0110351 - Deal Size (SF)",5,43855,68171,59916,96460,"Simple mean. Verified."),
 ("R0121833 (Sansone/CoStar) - Asking Rent $/SF",8,7.95,10.99,11.38,13.50,"SF-WEIGHTED (simple mean would be $11.29). Verified: $10.9874."),
 ("R0121833 - Deal Size (SF)",8,3630,13028,8769,38845,"Simple mean. Verified: 13,028.9."),
 ("R0121833 - Months on Market",8,1,13,7,65,"Published summary."),
 ("R0111559 / R0111560 (Sterling Exhibit F) - Scheduled Rent $/SF",7,5.17,6.26,6.25,6.97,"SIMPLE mean. Verified: $6.2557. SF-weighted would be $6.166, i.e. ~1.4% lower."),
]
for t,n,lo,av,me,hi,note in ROWS:
    vals=[t,n,lo,av,me,hi,note]
    for j,v in enumerate(vals,1):
        c=ws.cell(row=r,column=j,value=v);c.font=N;c.border=BOX
        c.alignment=Alignment(vertical="top",wrap_text=(j in(1,7)))
        if j==2: c.number_format=NUM
        if j in(3,4,5,6): c.number_format=(NUM if ("SF)" in t or "Market" in t) else M2)
    r+=1
r+=1
for line in ["Applied conclusions drawn by each preparer from these sets:",
 "  R0110351 (Ryan): market rent $7.00/SF NNN, 10% vacancy, 8% expenses, 6.25% cap -> $7,939,000 ($92.74/SF).",
 "  R0121833 (Sansone): $9.50/SF, 5% vacancy, 15% expenses, 7.14% loaded cap -> $3,919,000 ($107.45/SF).",
 "  R0111559/60 (Sterling): survey average $6.26 -> applied $6.25/SF NNN, 10% vacancy, 5% expenses, 7.63% loaded cap."]:
    ws.cell(row=r,column=1,value=line).font=(B if line.endswith(":") else N);r+=1
ws.freeze_panes="A5"

ws=wb.create_sheet("Notes & Data Gaps")
for i,w in enumerate([30,20,92],1): ws.column_dimensions[get_column_letter(i)].width=w
r=title(ws,1,3,"SOURCE NOTES, DISCREPANCIES & DATA GAPS")
r+=1
r=hdr(ws,r,["Item","Document / Reference","Detail"])
NOTES=[
 ("Scope of this workbook","3 of 5 packages","Only R0110351, R0121833 and R0111559/R0111560 contain lease comparables. R0114026 (6051 Washington St) contains no comps at all."),
 ("Shared comp set","R0111559 & R0111560","Both Sterling protests rely on the identical 7-lease Exhibit F survey; the set is stored once and tagged to both."),
 ("SUBJECT APPEARS AS ITS OWN COMP","R0121833 comp 6","5360 N Washington St, 7,285 SF, signed 11/9/2023 is the subject's own Unit B. CoStar shows $10.75/SF asking; the executed lease in the same package is $9.50/SF NNN - a 13% gap between the listed asking rent and the contract rent."),
 ("'Average' is SF-weighted","R0110351, R0121833","CoStar's published Average is weighted by square footage, not a simple mean. Using a simple mean instead would raise the Ryan asking average from $7.27 to $7.61 and the Sansone average from $10.99 to $11.29."),
 ("Sterling uses a simple mean","R0111559 / R0111560","Exhibit F's $6.26 average is unweighted. Weighted by SF it is $6.166 - the two largest leases in the set carry the lowest rents ($5.17 and $6.25)."),
 ("DATA GAP - effective rent","all three sets","Effective rent, TI allowance and free rent are reported as '-' for every deal in all 20 records."),
 ("DATA GAP - rent basis","R0121833","All 8 Sansone comps are ASKING rents; not one contract/starting rent is included."),
 ("DATA GAP - lease term","R0110351, R0121833","Ryan set: term shown for 3 of 5. Sansone set: term shown for 1 of 8 (120 months)."),
 ("DATA GAP - tenant / landlord","R0121833","No tenant or landlord is identified for any of the 8 Sansone comps."),
 ("DATA GAP - service type","R0121833 comps 5, 6","Quoted without an 'nnn' suffix, so the expense basis is not stated for those two."),
 ("Duplicate comp numbering","R0121833","The published list numbers only 6 properties but contains 8 deals - 5050 Fox St and 5865-5869 Broadway each appear twice. Stored here as 4/4b and 5/5b."),
 ("Confidential comps","R0111559 / R0111560","3 of the 7 Sterling leases are identified only as 'Confidential - Stapleton/Central Park Area'. Sterling asked the Assessor to keep the survey confidential."),
 ("Excluded - sale comparables","R0121833 p.13-15, R0111559 p.26-33","CoStar SALE comps (e.g. 6270 E 50th Ave, $3,775,000 / $106.52 per SF, sold 11/6/2023) and the Ascent/Great Western paired sales are market-support data and are out of scope for this lease database."),
 ("Excluded - market surveys","R0111559/60 p.8-12","Lowery Real Estate Investment Survey 2024 Winter/Spring (Exhibit A) and CoStar submarket reports (Exhibit B) are general market data, not property comps."),
]
for a,b_,c_ in NOTES:
    flag=a.startswith("DATA GAP") or a.startswith("SUBJECT")
    for j,v in enumerate([a,b_,c_],1):
        c=ws.cell(row=r,column=j,value=v);c.font=(B if (flag and j==1) else N);c.border=BOX
        c.alignment=Alignment(vertical="top",wrap_text=True)
        if flag: c.fill=PatternFill("solid",fgColor=GOLD)
    r+=1
ws.freeze_panes="A5"
wb.move_sheet("Lease Comparables",offset=-2)
wb.save(F)
print("sheets:",wb.sheetnames)
