# -*- coding: utf-8 -*-
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
from copy import copy
P="/home/user/Expense-Lease-Extractor-/Master_Property_Database.xlsx"
wb=load_workbook(P); NS="(not stated in document)"
MON='"$"#,##0.00'; MON0='"$"#,##0'; PSF='"$"#,##0.00'

def style_from(ws,r,nc): return [copy(ws.cell(row=r,column=c)._style) for c in range(1,nc+1)]
def put(ws,row,vals,st,nf=None):
    for j,v in enumerate(vals,1):
        c=ws.cell(row=row,column=j); c.value=v
        if j-1<len(st): c._style=copy(st[j-1])
        if nf and j in nf: c.number_format=nf[j]
        c.alignment=Alignment(vertical="top",wrap_text=True)
def snapshot(ws,frm,nc):
    return [[(ws.cell(row=r,column=c).value,copy(ws.cell(row=r,column=c)._style),ws.cell(row=r,column=c).number_format)
             for c in range(1,nc+1)] for r in range(frm,ws.max_row+1)]
def restore(ws,at,buf):
    for i,rd in enumerate(buf):
        for j,(v,s,f) in enumerate(rd,1):
            c=ws.cell(row=at+i,column=j); c.value=v; c._style=copy(s); c.number_format=f
def insert_block(ws,at,rows,src,nc,nf=None):
    n=len(rows)
    moved=[m for m in list(ws.merged_cells.ranges) if m.min_row>=at]
    for m in moved: ws.unmerge_cells(str(m))
    buf=snapshot(ws,at,nc); st=style_from(ws,src,nc)
    for i,r in enumerate(rows): put(ws,at+i,r,st,nf)
    restore(ws,at+n,buf)
    for m in moved:
        ws.merge_cells(start_row=m.min_row+n,start_column=m.min_col,end_row=m.max_row+n,end_column=m.max_col)

# ---------------- SOURCE DOCUMENTS
ws=wb["Source Documents"]
last=max(r for r in range(6,ws.max_row+1) if isinstance(ws.cell(row=r,column=1).value,(int,float)))
insert_block(ws,last+1,[
 [20,"R0187789.pdf","Assessor protest - income approach, recreated rent roll, 2yr expenses, 6 lease comps, 4 sale comps, LOA","5710 E. 56th Avenue, Commerce City (KEW Realty)","Tax Yr 2025 (val 6/30/2024)","Sterling Property Tax Specialists (Paul Leonard)"],
 [21,"R0188699.pdf","Tax appeal - market proforma + actual 2024 NNN income analysis + 3 lease comps + LOA","15100 East 40th Ave, Aurora - Confluent Center 70 (LBA Realty)","Tax Yr 2025 (val 6/30/2024)","Ryan LLC (Keegan Conway) / CoStar"],
 [22,"R0198179.pdf","Executed Second Amendment to Lease - State of Colorado (7-step rent schedule, TI allowance, renewal option)","17851 E 40th Avenue, Aurora","Renewal term 7/1/2026 - 6/30/2033","CO Dept of Personnel & Administration / Office of the State Architect"],
 [23,"R0191099.pdf","Assessor protest - income approach, 2 rent-roll dates, CY2024 income statement, base-period lease survey","3559 N Himalaya Rd, Bldg 5, Aurora (Majestic Commercenter Phase 9)","Tax Yr 2025 (val 6/30/2024)","Sterling Property Tax Specialists"],
 [24,"R0200721.pdf","Assessor protest - income approach, 3 rent-roll dates, CY2024 income statement, 5 base-period lease comps","3420 Lisbon Street, Aurora (Majestic Lisbon Buildings LLC)","Tax Yr 2025 (val 6/30/2024)","Sterling Property Tax Specialists"],
],6,6)

# ---------------- PROPERTY MASTER
ws=wb["Property Master"]; at=ws.max_row+1; st=style_from(ws,5,25)
pm=[
 ["R0187789","R0187789","R0187789","5710 E. 56th Avenue","5710 E. 56th Avenue","Commerce City","CO",NS,"Adams",
  "KEW Realty Corporation (David Spira, President)","KEW Realty Corporation","Industrial / Flex (multi-tenant)","Multi (3 tenants)",2.82,None,None,50000,2016,None,
  "100% (6/30/2024) - stabilized",6896593,"2025",5440000,"Sterling Property Tax Specialists / Goldstein Law Firm (agents)","R0187789.pdf"],
 ["R0188699","R0188699","R0188699","Confluent Center 70","15100 East 40th Avenue","Aurora","CO",NS,"Adams",
  "LBA Realty Fund","LBA NCC2-Company II, LLC","Industrial (30' clear height)","Multi/single not itemized",6.16,None,None,105174,2016,None,
  "100% (6/30/2024)",13272555,"2025",11498000,"LBA Realty (Stephanie Olazabal, VP Property Accounting)","R0188699.pdf"],
 ["R0198179","R0198179",NS,"17851 E 40th Avenue","17851 E 40th Avenue","Aurora","CO","80011","Adams",
  "Prologis, L.P. (successor-in-interest to DCT Summit LLC)","Prologis, L.P.","Industrial - leased to the State of Colorado","Single tenant (State of Colorado)",None,None,None,82131,None,None,
  "100% leased (State of Colorado)",None,NS,None,"Prologis, Inc. (general partner)","R0198179.pdf"],
 ["R0191099","R0191099","R0191099","Majestic Commercenter - Building 5","3559 N Himalaya Road","Aurora","CO","80011","Adams",
  "Majestic Commercenter Phase 9 (bldg 206605)","Majestic Commercenter Phase 9","Industrial - Warehouse","Multi (4 tenants)",None,None,None,159047,2017,None,
  "100% (6/30/2024 DOV)",19943316,"2025",11139000,"Majestic Realty Co. (Nicole Creighton, PM)","R0191099.pdf"],
 ["R0200721","R0200721","R0200721","3420 Lisbon Street","3420 N. Lisbon Street","Aurora","CO","80011","Adams",
  "Majestic Lisbon Buildings, LLC (bldg 347002)","Majestic Lisbon Buildings, LLC","Industrial - Warehouse","Single tenant",None,None,None,36212,2019,None,
  "100% (6/30/2024 DOV)",5477148,"2025",4463600,"Majestic Realty Co. (Nicole Creighton, PM)","R0200721.pdf"],
]
for i,r in enumerate(pm): put(ws,at+i,r,st,{14:'0.00',15:'#,##0',16:'#,##0',17:'#,##0',18:'0',21:MON0,23:MON0})

# ---------------- TENANTS & LEASES
ws=wb["Tenants & Leases"]; at=ws.max_row+1; st=style_from(ws,6,13)
tl=[
 ["5710 E. 56th Avenue (R0187789)","5710A","Prudential Overall Supply","Actual - NNN",10259.67,123116.00,12.31,"12/2017","4/2026","-",2091.67,None,
  "10,000 SF. Petitioner's recreated rent roll at 6/30/2024. Expense recoveries $25,100/yr ($2.51 PSF); no other income. Gross rent $14.82 PSF. Monthly figures derived from the annual amounts."],
 ["5710 E. 56th Avenue (R0187789)","5710C","United Rentals","Actual - NNN",8500.00,102000.00,10.20,"2/2019","4/2029","-",3650.00,None,
  "10,000 SF. Recoveries $43,800/yr ($4.38 PSF). Gross rent $14.58 PSF."],
 ["5710 E. 56th Avenue (R0187789)","5710H","Applied Control","Actual - NNN",31583.33,379000.00,12.63,"8/2018","12/2026","-",3325.00,None,
  "30,000 SF. Recoveries $39,900/yr ($1.33 PSF). Gross rent $13.96 PSF."],
 ["5710 E. 56th Avenue (R0187789)","TOTAL","3 tenants - 100% occupied","Actual - NNN",50343.00,604116.00,12.08,"-","-","-",9066.67,None,
  "50,000 SF occupied, 0 vacant. Printed totals: scheduled rent $604,116, expense recoveries $108,800, other income $0. Leases have 2 to 5 years remaining."],
 ["17851 E 40th Avenue (R0198179)","Whole","State of Colorado - Dept of Public Safety, Division of Homeland Security and Emergency Management","Actual - Gross (State-exempt from property tax)",62967.10,755605.20,9.20,"7/1/2026","6/30/2033","7-yr Second Renewal Term + one 5-yr Third Renewal Option",None,None,
  "82,131 RSF. Executed Second Amendment to a lease dated 12/22/2020 (First Amendment 3/15/2021; first renewal exercised 4/29/2021). Landlord Prologis, L.P. Figures shown are the first paying year. Full schedule (Adjusted Annual Rent PSF / Monthly): 8/1/26-6/30/27 $9.20 / $62,967.10; FY28 $9.47 / $64,815.05; FY29 $9.76 / $66,799.88; FY30 $10.05 / $68,784.71; FY31 $10.36 / $70,906.43; FY32 $10.67 / $73,028.15; FY33 $10.99 / $75,218.31. 7/1/26-7/31/26 is rent-free. Negotiated rent runs $11.52-$13.31 PSF before a $2.32 PSF property-tax credit. TI allowance $205,327.50 ($2.50 PSF); JLL tenant-agent commission $128,648.31."],
 ["Majestic Commercenter Bldg 5 (R0191099)","Suite 100","Erickson Metals of Colorado, Inc. (Randy Adkisson)","Actual - Multi-Tenant Net",29122.88,349474.56,6.7500,"12/1/2022","1/31/2033","-",None,None,
  "51,774 SF = 32.55% of building. Rent roll as of 12/31/23. This same lease is used as a comparable in Sterling's own Exhibit F survey for buildings 6 and 7, and in the R0200721 Exhibit C survey, where it is shown at $360,832 / $6.97 PSF rather than the $349,474.56 / $6.75 on this rent roll."],
 ["Majestic Commercenter Bldg 5 (R0191099)","-","Euromarket Designs, Inc.","Actual - Multi-Tenant Net",31693.10,380317.20,6.8790,"2/1/2018","3/31/2025","-",None,None,
  "55,287 SF = 34.76% of building. Rent roll as of 12/31/23."],
 ["Majestic Commercenter Bldg 5 (R0191099)","Suite 200","Carbon Chemistry Ltd.","Actual - Multi-Tenant Net",12260.40,147124.80,8.1955,"10/1/2020","1/31/2025","-",None,None,
  "17,952 SF = 11.29% of building. Rent roll as of 12/31/23."],
 ["Majestic Commercenter Bldg 5 (R0191099)","Suite 400","Service Partners, LLC (c/o TopBuild Corp)","Actual - Multi-Tenant Net",19838.25,238059.00,6.9947,"4/1/2021","6/30/2026","-",None,None,
  "34,034 SF = 21.40% of building. Rent roll as of 12/31/23; at 6/30/23 this lease was $19,260.43/mo, $231,125.16/yr, $6.7910 PSF."],
 ["Majestic Commercenter Bldg 5 (R0191099)","TOTAL 12/31/23","4 tenants - 100% occupied","Actual - Multi-Tenant Net",92914.63,1114975.56,7.0104,"-","-","-",None,None,
  "159,047 SF occupied, 0 vacant. Printed rent-roll totals at 12/31/23. At 6/30/23 the totals were $92,336.81/mo and $1,108,041.72/yr."],
 ["3420 Lisbon Street (R0200721)","Whole","Cameron Ashley Building Products, Inc. (Ryan Beaton)","Actual - Single Tenant Net",23370.53,280446.36,7.7446,"10/12/2021","12/31/2026","-",None,None,
  "36,212 SF, 100% of building. Rent roll as of 06/01/24. This annual rent ties exactly to Industrial Rental Income on the CY2024 income statement. Sterling's Exhibit C shows the same lease at $264,348 / $7.30 PSF scheduled with an effective rent of $5.86 PSF after 3 months free rent and $13.14 PSF of tenant improvements, while the letter text describes the commencement rate as $7.60 PSF - three different figures, all reproduced as printed."],
]
for i,r in enumerate(tl): put(ws,at+i,r,st,{5:MON,6:MON,7:PSF,11:MON,12:MON})

# ---------------- ASSESSMENTS
ws=wb["Assessments & Valuation"]
hdrB=[r for r in range(1,ws.max_row+1) if str(ws.cell(row=r,column=1).value or "").startswith("B. Income-Approach")][0]
lastB=max(r for r in range(hdrB+2,ws.max_row+1) if ws.cell(row=r,column=2).value is not None)
insert_block(ws,lastB+1,[
 ["5710 E. 56th Avenue - actual 2023 (R0187789)",640426,-51234,589192,-94857,494335,None,None],
 ["5710 E. 56th Avenue - actual 2024 (R0187789)",725742,-58059,667683,-107192,560491,None,None],
 ["5710 E. 56th Avenue - concluded, avg NOI (R0187789)",None,None,None,None,527413,0.0970,5440000],
 ["Confluent Center 70 - market proforma (R0188699)",815099,-65208,749891,-59991,689899,0.0600,11498000],
 ["Confluent Center 70 - actual 2024 NNN (R0188699)",1536047,None,1536047,-782165,753882,None,None],
 ["Majestic Commercenter Bldg 5 (R0191099)",994044,-99404,894639,-44732,849907,0.0763,11139000],
 ["3420 Lisbon Street (R0200721)",398332,-39833,358499,-17925,340574,0.0763,4463600],
],hdrB+2,8,{2:MON0,3:MON0,4:MON0,5:MON0,6:MON0,7:'0.00%',8:MON0})
lastA=max(r for r in range(6,hdrB) if ws.cell(row=r,column=2).value is not None)
insert_block(ws,lastA+1,[
 ["5710 E. 56th Avenue (R0187789)","2025",6896593,137.93,5440000,108.80,None,"R0187789"],
 ["Confluent Center 70 (R0188699)","2025",13272555,126.20,11498000,109.32,12707001,"R0188699"],
 ["17851 E 40th Avenue (R0198179)",NS,None,None,None,None,None,"R0198179 - lease document only, no valuation"],
 ["Majestic Commercenter Bldg 5 (R0191099)","2025",19943316,125.39,11139000,70.04,None,"R0191099"],
 ["3420 Lisbon Street (R0200721)","2025",5477148,151.25,4463600,123.26,None,"R0200721"],
],6,8,{3:MON0,4:PSF,5:MON0,6:PSF,7:MON0})

# ---------------- I&E SUMMARY
ws=wb["Income-Expense Summary"]
at=max(r for r in range(5,ws.max_row+1) if ws.cell(row=r,column=1).value and ws.cell(row=r,column=5).value)+1
ies=[
 ["5710 E. 56th Avenue - CY2023",640426,214395,"Owner-reported. Total expenses INCLUDE real estate taxes ($137,214); the petitioner's NOI of $494,335 strips taxes out, adds a 3% reserve and applies 8% vacancy. R0187789"],
 ["5710 E. 56th Avenue - CY2024",725742,228330,"Owner-reported. Includes real estate taxes ($141,168); petitioner's NOI $560,491 on the same basis as 2023. Two-year average NOI $527,413. R0187789"],
 ["Confluent Center 70 - CY2024",1536047,782165,"Actual NNN, accrual. Expenses include property taxes $391,900. TRUE NOI $753,882 ($7.17 PSF) - no debt service or depreciation. R0188699"],
 ["Majestic Commercenter Bldg 5 - CY2024",1800360.37,628819.46,"Operating income before interest ($331,694.13) and other/depreciation ($559,919.27); net income after those = $279,927.51. R0191099"],
 ["3420 Lisbon Street - CY2024",484482.38,205776.75,"Operating income before interest ($86,641.63) and other/depreciation ($233,804.17); net income after those = ($41,740.17). R0200721"],
]
fixed=[]
for i,r in enumerate(ies):
    row=at+i; fixed.append([r[0],r[1],r[2],f"=B{row}-C{row}",r[3]])
insert_block(ws,at,fixed,5,5,{2:MON,3:MON,4:MON})
wb.save(P); print("part1 saved; I&E summary at row",at)
