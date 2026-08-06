# -*- coding: utf-8 -*-
from openpyxl import Workbook
from openpyxl.styles import Font,PatternFill,Alignment,Border,Side
from openpyxl.utils import get_column_letter
from datetime import date
import statistics as st

NAVY="FF1F3864";BLUE="FF2E5C8A";GREEN="FFE2EFDA";GOLD="FFFFF2CC";GREY="FFF2F2F2"
W=Font(name="Arial",size=10,color="FFFFFFFF",bold=True);TT=Font(name="Arial",size=12,color="FFFFFFFF",bold=True)
B=Font(name="Arial",size=10,bold=True);N=Font(name="Arial",size=10);I=Font(name="Arial",size=9,italic=True,color="FF595959")
th=Side(style="thin",color="FFBFBFBF");BOX=Border(left=th,right=th,top=th,bottom=th)
M2='"$"#,##0.00';NUM='#,##0';DT='mm/dd/yyyy';MON='"$"#,##0'
wb=Workbook()
def title(ws,r,nc,t,sub=None):
    ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=nc)
    c=ws.cell(row=r,column=1,value=t);c.font=TT;c.fill=PatternFill("solid",fgColor=NAVY);ws.row_dimensions[r].height=22
    if sub:
        ws.merge_cells(start_row=r+1,start_column=1,end_row=r+1,end_column=nc)
        ws.cell(row=r+1,column=1,value=sub).font=I;return r+2
    return r+1
def hdr(ws,r,v):
    for j,x in enumerate(v,1):
        c=ws.cell(row=r,column=j,value=x);c.font=W;c.fill=PatternFill("solid",fgColor=BLUE)
        c.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True);c.border=BOX
    ws.row_dimensions[r].height=34;return r+1
def widths(ws,w):
    for i,x in enumerate(w,1): ws.column_dimensions[get_column_letter(i)].width=x

# ============================================================== ALL COMPS
ws=wb.active;ws.title="Lease Comparables"
cols=["Set","Comp #","Property / Building","Address","City","ST","ZIP","Submarket / County","SF Leased",
 "Sign Date","Lease Start","Lease Expire","Term","Rent $/SF","Rent Basis","Service","Deal Type","Lease Type",
 "Annual Rent $","Tenant","Landlord","Year Built","Source (PDF)"]
widths(ws,[26,7,30,30,10,5,8,20,11,12,12,12,13,10,10,9,11,13,14,26,26,10,14])
r=title(ws,1,len(cols),"LEASE COMPARABLES - 5 ADAMS COUNTY APPEAL PACKAGES",
  "20 comparable lease records from 3 of the 5 documents. R0111559 and R0111560 rely on the same 7-lease survey.")
r+=1;hr=r;r=hdr(ws,r,cols)
first=r
C=[]
# --- Ryan / CoStar set for R0110351
S1="R0110351 - 18300 E 28th Ave"
C+=[
 [S1,1,"Majestic Commercenter - Bldg 22","3503-3593 N Windsor Dr","Aurora","CO","80011","SW DIA/Pena Blvd",85253,date(2023,10,12),"Nov 2023",None,"1 Year",6.34,"Asking","NNN","New Sublease","New Lease",None,"Aspen Distribution","Majestic Realty Co.",None,"R0110351 p.4-6"],
 [S1,2,"Building 8","4555-4685 Geneva St","Denver","CO","80238","East I-70/270",96460,date(2023,8,7),"Feb 2024",None,"5 Years 1 Month",7.35,"Starting","NNN","Renewal Direct","Renewal",None,"Forward","Prologis, Inc.",None,"R0110351 p.4-6"],
 [S1,3,"Bldg 1","13550 E Smith Rd","Aurora","CO","80011","SW DIA/Pena Blvd",43855,date(2023,3,21),"Aug 2023",None,None,9.75,"Asking","NNN","New Direct","New Lease",None,"AcoustiFlo","Lincoln Property Company",None,"R0110351 p.4-6"],
 [S1,4,"Bldg 24","3500 N Windsor Dr","Aurora","CO","80011","SW DIA/Pena Blvd",59916,date(2022,11,29),"Dec 2022",None,"2 Years",6.50,"Starting","NNN","Renewal Direct","Renewal",None,"Anderson Corporation","Majestic Realty Co.",None,"R0110351 p.4-6"],
 [S1,5,"Bldg 5","2255 N Pagosa St","Aurora","CO","80011","SW DIA/Pena Blvd",55375,date(2022,8,14),"Jan 2023",None,None,6.75,"Asking","NNN","New Direct","New Lease",None,None,"EastGroup Properties, Inc.",None,"R0110351 p.4-6"],
]
# --- Sansone / CoStar set for R0121833
S2="R0121833 - Anchor Business Park"
C+=[
 [S2,1,"5005 Washington St","5005 Washington St","Denver","CO",None,"Adams / Denver north",38845,date(2024,5,8),None,None,None,11.50,"Asking","NNN","Direct","New Lease",None,None,None,None,"R0121833 p.11-12"],
 [S2,2,"5635 Franklin St","5635 Franklin St","Denver","CO",None,"Adams / Denver north",7500,date(2024,4,28),None,None,None,13.50,"Asking","NNN","Direct","New Lease",None,None,None,None,"R0121833 p.11-12"],
 [S2,3,"Bldg D","4930 Fox St","Denver","CO",None,"Adams / Denver north",8400,date(2024,4,8),None,None,None,10.95,"Asking","NNN","Direct","New Lease",None,None,None,None,"R0121833 p.11-12"],
 [S2,4,"5050 Fox St","5050 Fox St","Denver","CO",None,"Adams / Denver north",11200,date(2024,2,15),None,None,None,11.95,"Asking","NNN","Direct","New Lease",None,None,None,None,"R0121833 p.11-12"],
 [S2,"4b","5050 Fox St","5050 Fox St","Denver","CO",None,"Adams / Denver north",3630,date(2024,2,15),None,None,None,12.50,"Asking","NNN","Direct","New Lease",None,None,None,None,"R0121833 p.11-12"],
 [S2,5,"5865-5869 Broadway","5865-5869 Broadway","Denver","CO",None,"Adams / Denver north",9138,date(2024,1,3),None,None,None,11.25,"Asking","(not stated)","Direct","New Lease",None,None,None,None,"R0121833 p.11-12"],
 [S2,"5b","5865-5869 Broadway","5865-5869 Broadway","Denver","CO",None,"Adams / Denver north",18233,date(2023,12,1),None,None,None,7.95,"Asking","NNN","Direct","New Lease",None,None,None,None,"R0121833 p.11-12"],
 [S2,6,"5360 N Washington St (THE SUBJECT)","5360 N Washington St","Denver","CO","80216","Adams / Denver north",7285,date(2023,11,9),None,None,"120.0 months",10.75,"Asking","(not stated)","Direct","New Lease",None,"Integrative Environmental Systems, LLC (per lease)","Anchor Business Park, L.L.C.",2001,"R0121833 p.11-12"],
]
# --- Sterling base-period lease survey (Exhibit F), used by BOTH R0111559 and R0111560
S3="R0111559 / R0111560 - Exhibit F"
SB=[("3559 N. Himalaya Road, Suite 100","Aurora","Adams",date(2022,12,1),date(2033,1,31),10.2,51774,360832,6.97,2023),
    ("19682 E. 34th Drive","Aurora","Adams",date(2024,5,1),date(2027,4,30),3.0,200002,1033825,5.17,1997),
    ("3700 N. Windsor Drive","Aurora","Adams",date(2022,6,1),date(2032,5,31),10.0,210195,1313719,6.25,2013),
    ("Confidential - Stapleton/Central Park Area","Denver","Denver",date(2022,10,1),date(2027,9,30),5.0,115829,712344,6.15,2014),
    ("4735 Florence Street","Denver","Denver",date(2024,1,18),date(2029,12,31),6.0,247654,1646899,6.65,2002),
    ("Confidential - Stapleton/Central Park Area","Denver","Denver",date(2022,7,1),date(2027,6,30),5.0,80623,511956,6.35,2014),
    ("Confidential - Stapleton/Central Park Area","Denver","Denver",date(2022,6,1),date(2029,8,31),7.3,81731,510819,6.25,2014)]
for i,(ad,ci,co,s,e,yr,sf,ann,psf,yoc) in enumerate(SB,1):
    C.append([S3,i,ad,ad,ci,"CO",None,co+" County",sf,None,s,e,f"{yr} years",psf,"Scheduled/contract","NNN","Direct",None,ann,None,None,yoc,"R0111559 p.34 / R0111560 p.35"])

for row in C:
    for j,v in enumerate(row,1):
        c=ws.cell(row=r,column=j,value=v);c.font=N;c.border=BOX
        c.alignment=Alignment(vertical="top",wrap_text=(j in(3,4,20,21)))
        if j in(10,11,12) and isinstance(v,date): c.number_format=DT
        if j==9: c.number_format=NUM
        if j==14: c.number_format=M2
        if j==19: c.number_format=MON
        if j==22: c.number_format='0'
    if row[0]==S2:
        for j in range(1,len(cols)+1): ws.cell(row=r,column=j).fill=PatternFill("solid",fgColor=GREY)
    if "SUBJECT" in str(row[2]):
        for j in range(1,len(cols)+1): ws.cell(row=r,column=j).fill=PatternFill("solid",fgColor=GOLD)
    r+=1
last=r-1
ws.freeze_panes=f"C{first}";ws.auto_filter.ref=f"A{hr}:{get_column_letter(len(cols))}{last}"
r+=1
ws.cell(row=r,column=1,value="Highlighted row = comp 6 in the Sansone set is the subject property's own Unit B space. CoStar lists it at $10.75/SF asking; the executed lease in the same package is $9.50/SF NNN.").font=I
wb.save("Lease_Comparables_5_Adams_County_Packages.xlsx")
print("saved rows:",len(C))

# ---- verification of published stats
ryan=[c for c in C if c[0]==S1]; sans=[c for c in C if c[0]==S2]; ster=[c for c in C if c[0]==S3]
def wavg(rows): return sum(c[13]*c[8] for c in rows)/sum(c[8] for c in rows)
ask=[c for c in ryan if c[14]=="Asking"]; sta=[c for c in ryan if c[14]=="Starting"]
print()
print("RYAN set - asking : n=%d low %.2f wavg %.4f med %.2f high %.2f  (doc: 3 / 6.34 / 7.27 / 6.75 / 9.75)"%(len(ask),min(c[13] for c in ask),wavg(ask),st.median([c[13] for c in ask]),max(c[13] for c in ask)))
print("RYAN set - starting: n=%d low %.2f wavg %.4f med %.2f high %.2f  (doc: 2 / 6.50 / 7.02 / 6.93 / 7.35)"%(len(sta),min(c[13] for c in sta),wavg(sta),st.median([c[13] for c in sta]),max(c[13] for c in sta)))
print("SANSONE  - asking : n=%d low %.2f wavg %.4f med %.3f high %.2f  (doc: 8 / 7.95 / 10.99 / 11.38 / 13.50)"%(len(sans),min(c[13] for c in sans),wavg(sans),st.median([c[13] for c in sans]),max(c[13] for c in sans)))
print("SANSONE  - size   : avg %.1f med %.0f low %d high %d  (doc: 13,028 / 8,769 / 3,630 / 38,845)"%(st.mean([c[8] for c in sans]),st.median([c[8] for c in sans]),min(c[8] for c in sans),max(c[8] for c in sans)))
print("STERLING - simple avg %.4f  median %.2f  (doc: 6.26 / 6.25);  SF-weighted would be %.3f"%(st.mean([c[13] for c in ster]),st.median([c[13] for c in ster]),sum(c[18] for c in ster)/sum(c[8] for c in ster)))
print("STERLING - per-lease PSF recomputed from annual/SF:",[round(c[18]/c[8],2) for c in ster])
