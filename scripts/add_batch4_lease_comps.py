# -*- coding: utf-8 -*-
from openpyxl import load_workbook
from openpyxl.styles import Font,PatternFill,Alignment,Border,Side
from openpyxl.utils import get_column_letter
from datetime import date
from copy import copy
import statistics as st
F="LC2.xlsx"; wb=load_workbook(F)
N=Font(name="Arial",size=10); I=Font(name="Arial",size=9,italic=True,color="FF595959"); B=Font(name="Arial",size=10,bold=True)
th=Side(style="thin",color="FFBFBFBF"); BOX=Border(left=th,right=th,top=th,bottom=th)
M2='"$"#,##0.00'; NUM='#,##0'; DT='mm/dd/yyyy'; MON='"$"#,##0'; GOLD="FFFFF2CC"
ws=wb["Lease Comparables"]
last=max(r for r in range(5,ws.max_row+1) if ws.cell(row=r,column=2).value not in (None,""))
S="R0164588 - 9410 Heinz Way"
NEW=[
 [S,1,"4744 Forest St","4744 Forest St","Denver","CO","80216","East I-70/270",71060,date(2024,11,14),date(2024,11,1),None,"3 Years 2 Months",6.35,"Starting","NNN","Direct","New Lease",None,"Rugby Holdings, LLC","Travelers",None,"R0164588 p.4-6"],
 [S,2,"Upland III","3250 Abilene St","Aurora","CO","80011","SW DIA/Pena Blvd",80819,date(2024,11,5),date(2025,4,1),None,None,6.95,"Asking","NNN","Direct","New Lease",None,"Flexpax","Blackstone Real Estate Income Trust",None,"R0164588 p.4-6"],
 [S,3,"Building 2","11551 E 45th Ave","Denver","CO","80239","Cent E I-70/Montbello",62614,date(2024,7,24),date(2024,12,1),None,None,6.95,"Asking","NNN","Direct","New Lease",None,"Colorado Distribution Group","Clarion Partners",None,"R0164588 p.4-6"],
 [S,4,"4221 Monaco St","4221 Monaco St","Denver","CO","80216","Quebec St",265361,date(2023,4,18),date(2023,4,1),None,None,5.95,"Asking","NNN","Direct","New Lease",None,"RFMX, Corp / Houger Express LLC","Conscience Bay Company",None,"R0164588 p.4-6"],
 [S,5,"Colorado Logistics Park - Building D","10899 Havana St","Commerce City","CO","80640","DIA",82744,date(2022,8,4),date(2023,2,1),None,None,5.45,"Starting","NNN","Direct","New Lease",None,"Würth","Brennan Investment Group",None,"R0164588 p.4-6"],
]
src=[copy(ws.cell(row=last,column=c)._style) for c in range(1,24)]
noterows=[]
for r in range(last+1,ws.max_row+1):
    v=ws.cell(row=r,column=1).value
    if v: noterows.append(v); ws.cell(row=r,column=1).value=None
at=last+1
for i,row in enumerate(NEW):
    for j,v in enumerate(row,1):
        c=ws.cell(row=at+i,column=j); c.value=v; c._style=copy(src[j-1]); c.border=BOX; c.font=N
        c.alignment=Alignment(vertical="top",wrap_text=(j in(3,4,20,21,23)))
        if j in(10,11,12) and isinstance(v,date): c.number_format=DT
        if j==9: c.number_format=NUM
        if j==14: c.number_format=M2
        if j==19: c.number_format=MON
        if j==22: c.number_format='0'
        c.fill=PatternFill("solid",fgColor="FFFFFFFF")
r=at+len(NEW)+1
for t in noterows: ws.cell(row=r,column=1,value=t).font=I; r+=1
ws.auto_filter.ref=f"A4:{get_column_letter(23)}{at+len(NEW)-1}"

ws=wb["Set Statistics"]
tgt=max(r for r in range(5,ws.max_row+1) if isinstance(ws.cell(row=r,column=2).value,(int,float)))
src=[copy(ws.cell(row=tgt,column=c)._style) for c in range(1,8)]
buf=[[(ws.cell(row=rr,column=c).value,copy(ws.cell(row=rr,column=c)._style),ws.cell(row=rr,column=c).number_format)
      for c in range(1,8)] for rr in range(tgt+1,ws.max_row+1)]
ask=[6.95,6.95,5.95]; asf=[80819,62614,265361]; sta=[6.35,5.45]; ssf=[71060,82744]
new=[["R0164588 (Ryan/CoStar) - Asking Rent $/SF",3,5.95,6.30,6.95,6.95,
      "As published. SF-WEIGHTED - verified at $6.3005. A simple mean would be $6.62."],
     ["R0164588 (Ryan/CoStar) - Starting Rent $/SF",2,5.45,5.87,5.90,6.35,
      "As published. SF-weighted, verified at $5.8658. The 'median' column ($5.90) is the simple mean of the two deals."]]
for i,row in enumerate(new):
    for j,v in enumerate(row,1):
        c=ws.cell(row=tgt+1+i,column=j); c.value=v; c._style=copy(src[j-1])
        c.alignment=Alignment(vertical="top",wrap_text=(j in(1,7)))
        if j==2: c.number_format=NUM
        if j in(3,4,5,6): c.number_format=M2
for i,rd in enumerate(buf):
    for j,(v,s,f) in enumerate(rd,1):
        c=ws.cell(row=tgt+1+len(new)+i,column=j); c.value=v; c._style=copy(s); c.number_format=f

ws=wb["Notes & Data Gaps"]
last=max(r for r in range(5,ws.max_row+1) if ws.cell(row=r,column=1).value)
src=[copy(ws.cell(row=last,column=c)._style) for c in range(1,4)]
add=[("Scope update - 20 packages","batch 4","Of the 5 newest packages only R0164588 carries a new comp set (5 comps). R0187787 reuses the identical 6-comp set already stored for R0187789. R0157664, R0164287 and R0180894 contain no comparables."),
 ("Identical comp set across siblings","R0187787 / R0187789","Sterling used the same six CoStar comps, in the same order and at the same rents, for both 5690 and 5710 E. 56th Avenue - adjacent KEW Realty buildings protested by the same analyst on the same day. The set is stored once, tagged to R0187789."),
 ("Richer detail available","R0187787 p.14-16","The Exhibit C write-ups add asking-vs-starting rent, term, expiry, months on market, build-out, landlord and leasing rep for each of the six comps - e.g. comp 1 (6863-6865 E 48th Ave) asks $8.95/NNN but started at $9.05/NNN on a 5-year term expiring Apr 2028."),
 ("DATA GAP - term and concessions","R0164588","Term is reported for 1 of the 5 comps (38 months); free rent, TI allowance and effective rent are blank for all five."),
]
for i,(a_,b_,c_) in enumerate(add):
    flag=a_.startswith("DATA GAP")
    for j,v in enumerate([a_,b_,c_],1):
        c=ws.cell(row=last+1+i,column=j); c.value=v; c._style=copy(src[j-1])
        c.font=B if (flag and j==1) else N; c.border=BOX
        c.alignment=Alignment(vertical="top",wrap_text=True)
        if flag: c.fill=PatternFill("solid",fgColor=GOLD)
wb.save(F)
print("asking  weighted %.4f (doc 6.30) simple %.4f median %.2f"%(sum(a*b for a,b in zip(ask,asf))/sum(asf),st.mean(ask),st.median(ask)))
print("starting weighted %.4f (doc 5.87) simple %.4f"%(sum(a*b for a,b in zip(sta,ssf))/sum(ssf),st.mean(sta)))
w=load_workbook(F); print("comp rows:",sum(1 for r in range(5,w['Lease Comparables'].max_row+1) if w['Lease Comparables'].cell(row=r,column=2).value not in (None,"")))
