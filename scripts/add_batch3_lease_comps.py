# -*- coding: utf-8 -*-
from openpyxl import load_workbook
from openpyxl.styles import Font,PatternFill,Alignment,Border,Side
from openpyxl.utils import get_column_letter
from datetime import date
from copy import copy
import statistics as st
F="LC.xlsx"; wb=load_workbook(F)
N=Font(name="Arial",size=10); I=Font(name="Arial",size=9,italic=True,color="FF595959"); B=Font(name="Arial",size=10,bold=True)
th=Side(style="thin",color="FFBFBFBF"); BOX=Border(left=th,right=th,top=th,bottom=th)
M2='"$"#,##0.00'; NUM='#,##0'; DT='mm/dd/yyyy'; MON='"$"#,##0'; GOLD="FFFFF2CC"; GREY="FFF2F2F2"
ws=wb["Lease Comparables"]
last=max(r for r in range(5,ws.max_row+1) if ws.cell(row=r,column=2).value not in (None,""))
S1="R0187789 - 5710 E. 56th Ave"; S2="R0188699 - Confluent Center 70"; S3="R0200721 - 3420 Lisbon St"
NEW=[
 # --- R0187789 (Sterling / CoStar), 6 comps
 [S1,1,"6863-6865 E 48th Ave","6863-6865 E 48th Ave","Denver","CO",None,"Adams / Denver north",7300,None,date(2023,1,1),None,None,9.05,"Contract","NNN","Direct",None,None,"Raul's Boat & RV Upholstery",None,None,"R0187789 p.2"],
 [S1,2,"6900 E 47th Avenue Dr","6900 E 47th Avenue Dr","Denver","CO",None,"Adams / Denver north",6830,None,date(2023,12,1),None,None,9.25,"Contract","NNN","Direct",None,None,"Prime Controls LP",None,None,"R0187789 p.2 (Flex)"],
 [S1,3,"6660-6722 E 47th Avenue Dr","6660-6722 E 47th Avenue Dr","Denver","CO",None,"Adams / Denver north",4000,None,date(2023,8,1),None,None,9.50,"Contract","NNN","Direct",None,None,"Midwest Connect, LLC",None,None,"R0187789 p.2"],
 [S1,4,"6804 E 48th Ave","6804 E 48th Ave","Denver","CO",None,"Adams / Denver north",5808,None,date(2023,6,1),None,None,10.00,"Contract","NNN","Direct",None,None,"Logistics Union, Inc",None,None,"R0187789 p.2"],
 [S1,5,"6751-6785 E 50th Ave","6751-6785 E 50th Ave","Commerce City","CO",None,"Adams",13200,None,date(2024,2,1),None,None,10.19,"Contract","NNN","Direct",None,None,"Aqua Rocks Colorado",None,None,"R0187789 p.2"],
 [S1,6,"3975 E 56th Ave","3975 E 56th Ave","Commerce City","CO",None,"Adams",1250,None,date(2023,6,1),None,None,10.95,"Contract","NNN","Direct",None,None,"Sunshine Home Services",None,None,"R0187789 p.2"],
 # --- R0188699 (Ryan / CoStar), 3 comps
 [S2,1,"Building 8","4555-4685 Geneva St & E 47th Ave","Denver","CO","80238",None,96460,date(2023,8,1),None,None,"5.1 years",7.35,"Starting","NNN","Direct","New Lease",None,None,None,None,"R0188699 p.5"],
 [S2,2,"Ascent Commerce Center - Building 3","18146 E 84th Ave","Commerce City","CO",None,None,26463,date(2023,8,1),None,None,"5.0 years",9.50,"Starting","NNN","Direct","New Lease",None,None,None,None,"R0188699 p.5"],
 [S2,3,"Bldg 24","3500 N Windsor Dr","Aurora","CO","80011","SW DIA/Pena Blvd",59916,date(2022,11,1),None,None,"2.0 years",6.50,"Starting","NNN","Direct","New Lease",None,None,None,None,"R0188699 p.5"],
]
# --- R0200721 Exhibit C (Sterling base-period survey), 5 comps + subject prior-period lease
EX=[("3250 N. Himalaya Road, Suite 200",date(2024,6,1),date(2029,7,31),"5.2 years",33056,330560,10.00,7.81,2023,"Multi","2 months free rent; tenant TIs $22.39/SF. Free-rent loss $10,662/yr ($0.32 PSF); TI adjustment $61,667/yr ($1.87 PSF)."),
    ("19655 East 35th Drive, Suite 700",date(2022,8,1),date(2027,3,31),"4.7 years",48252,345506,7.16,7.16,2007,"Multi","No free rent, no TI adjustment - scheduled equals effective."),
    ("3559 N. Himalaya Road, Suite 100",date(2022,12,1),date(2033,1,31),"10.2 years",51774,360832,6.97,6.97,2023,"Multi","No free rent or TI adjustment. This is the Erickson Metals lease AT subject property R0191099 - see Notes tab."),
    ("17654 Spinnaker Way, Suite 100",date(2024,6,1),date(2029,8,31),"5.3 years",20997,248814,11.85,11.19,2024,"Multi","3 months free rent; tenant TIs $1.19/SF. Free-rent loss $11,844/yr ($0.56 PSF); TI adjustment $2,082/yr ($0.10 PSF)."),
    ("17554 Spinnaker Way, Suite 2",date(2024,1,30),date(2031,4,30),"7.3 years",48808,512484,10.50,8.93,2024,"Multi","3 months free rent; tenant TIs $14.50/SF. Free-rent loss $17,667/yr ($0.36 PSF); TI adjustment $58,988/yr ($1.21 PSF).")]
for i,(ad,s,e,tm,sf,ann,psf,eff,yoc,mt,note) in enumerate(EX,1):
    NEW.append([S3,i,ad,ad,"Aurora","CO",None,"Adams",sf,None,s,e,tm,psf,f"Scheduled (effective ${eff:.2f})","NNN","Direct",None,ann,None,None,yoc,"R0200721 p.16 - "+note])
NEW.append([S3,"subj","3420 Lisbon St. - THE SUBJECT (prior base period)","3420 N. Lisbon Street","Aurora","CO","80011","Adams",36212,None,date(2021,10,1),date(2026,12,31),"5.3 years",7.30,"Scheduled (effective $5.86)","NNN","Direct",None,264348,"Cameron Ashley Building Products, Inc.","Majestic Lisbon Buildings, LLC",2019,
  "R0200721 p.16 - subject's own prior-base-period lease, listed at the bottom of Sterling's survey. 3 months free rent; tenant TIs $13.14/SF. The 06/01/24 rent roll shows this same lease at $280,446.36 / $7.7446 PSF and the letter text describes it as $7.60 PSF."])
src=[copy(ws.cell(row=last,column=c)._style) for c in range(1,24)]
noterows=[]
for r in range(last+1,ws.max_row+1):
    v=ws.cell(row=r,column=1).value
    if v: noterows.append(v); ws.cell(row=r,column=1).value=None
at=last+1
for i,row in enumerate(NEW):
    for j,v in enumerate(row,1):
        c=ws.cell(row=at+i,column=j); c.value=v; c._style=copy(src[j-1]); c.border=BOX; c.font=N
        c.alignment=Alignment(vertical="top",wrap_text=(j in(3,4,20,21,23)))
        if j in(10,11,12) and isinstance(v,date): c.number_format=DT
        if j==9: c.number_format=NUM
        if j==14: c.number_format=M2
        if j==19: c.number_format=MON
        if j==22: c.number_format='0'
        c.fill=PatternFill("solid",fgColor=(GOLD if "SUBJECT" in str(row[2]) else ("FFFFFFFF" if row[0]!=S2 else GREY)))
r=at+len(NEW)+1
for t in noterows: ws.cell(row=r,column=1,value=t).font=I; r+=1
ws.auto_filter.ref=f"A4:{get_column_letter(23)}{at+len(NEW)-1}"

# ---- Set Statistics
ws=wb["Set Statistics"]
tgt=max(r for r in range(5,ws.max_row+1) if isinstance(ws.cell(row=r,column=2).value,(int,float)))
src=[copy(ws.cell(row=tgt,column=c)._style) for c in range(1,8)]
buf=[[(ws.cell(row=rr,column=c).value,copy(ws.cell(row=rr,column=c)._style),ws.cell(row=rr,column=c).number_format)
      for c in range(1,8)] for rr in range(tgt+1,ws.max_row+1)]
a=[9.05,9.25,9.50,10.00,10.19,10.95]; asf=[7300,6830,4000,5808,13200,1250]
b=[7.35,9.50,6.50]; bsf=[96460,26463,59916]
c5=[10.00,7.16,6.97,11.85,10.50]; e5=[7.81,7.16,6.97,11.19,8.93]; c5sf=[33056,48252,51774,20997,48808]
new=[
 ["R0187789 (Sterling/CoStar) - Rent $/SF",len(a),min(a),round(st.mean(a),2),st.median(a),max(a),
  "No summary block in the source; computed here. Simple mean $%.2f; SF-weighted $%.2f."%(st.mean(a),sum(x*y for x,y in zip(a,asf))/sum(asf))],
 ["R0188699 (Ryan/CoStar) - Starting Rent $/SF",3,6.50,7.78,7.35,9.50,
  "As published. Verified: simple mean $7.7833. The source separately prints a SF-weighted figure of $7.38 - verified at $7.3826."],
 ["R0200721 (Sterling Exhibit C) - Scheduled Rent $/SF",5,6.97,9.30,10.00,11.85,
  "As published. Simple mean verified at $9.296. SF-weighted would be $9.06."],
 ["R0200721 (Sterling Exhibit C) - Effective Rent $/SF",5,6.97,8.41,7.81,11.19,
  "As published, after free rent and TI amortisation. Simple mean verified at $8.412."],
]
for i,row in enumerate(new):
    for j,v in enumerate(row,1):
        c=ws.cell(row=tgt+1+i,column=j); c.value=v; c._style=copy(src[j-1])
        c.alignment=Alignment(vertical="top",wrap_text=(j in(1,7)))
        if j==2: c.number_format=NUM
        if j in(3,4,5,6): c.number_format=M2
for i,rd in enumerate(buf):
    for j,(v,s,f) in enumerate(rd,1):
        c=ws.cell(row=tgt+1+len(new)+i,column=j); c.value=v; c._style=copy(s); c.number_format=f

# ---- Notes
ws=wb["Notes & Data Gaps"]
last=max(r for r in range(5,ws.max_row+1) if ws.cell(row=r,column=1).value)
src=[copy(ws.cell(row=last,column=c)._style) for c in range(1,4)]
add=[("Scope update - 15 packages","batch 3","Of the 5 newest packages, 3 carry lease comparables (R0187789, R0188699, R0200721). R0191099 relies on the same 7-lease Exhibit F survey already stored for R0111559/R0111560, and R0198179 is an executed lease, not a comp set."),
 ("SUBJECT APPEARS AS ITS OWN COMP","R0200721 Exhibit C","3420 Lisbon St. is listed at the foot of its own base-period survey as a 'Prior Base period' lease at $7.30/SF scheduled, $5.86/SF effective. Its own 06/01/24 rent roll shows $7.7446/SF and the letter text says $7.60/SF - three different figures for one lease."),
 ("CROSS-PORTFOLIO SELF-COMP","R0200721 comp 3 / R0191099","'3559 N. Himalaya Road, Suite 100' in the Lisbon survey is the Erickson Metals lease at subject property R0191099, and the same lease also anchors the Exhibit F survey used for buildings 6 and 7. Sterling shows it at $360,832 / $6.97 PSF in the surveys but the R0191099 rent roll carries it at $349,474.56 / $6.75 PSF."),
 ("Petitioner applied ABOVE its own survey","R0200721","Sterling's survey averages $9.30/SF scheduled and $8.41/SF effective, yet the income analysis applies $11.00/SF - higher than every comp except 17654 Spinnaker Way."),
 ("Repeat comps across packages","R0188699 comps 1 & 3","'Building 8' (4555-4685 Geneva St) and 'Bldg 24' (3500 N Windsor Dr) are the same two transactions already captured in the R0110351 comp set, at the same $7.35 and $6.50 starting rents."),
 ("DATA GAP - sign dates","R0187789","The 6 comps report a lease start month only; no sign date, term, deal type, landlord or building size is given."),
 ("DATA GAP - concessions","R0187789, R0188699","Free rent, TI allowance and effective rent are blank for all 9 comps in these two sets. Only the R0200721 survey quantifies concessions."),
]
for i,(a_,b_,c_) in enumerate(add):
    flag=a_.startswith("DATA GAP") or a_.startswith("SUBJECT") or a_.startswith("CROSS")
    for j,v in enumerate([a_,b_,c_],1):
        c=ws.cell(row=last+1+i,column=j); c.value=v; c._style=copy(src[j-1])
        c.font=B if (flag and j==1) else N; c.border=BOX
        c.alignment=Alignment(vertical="top",wrap_text=True)
        if flag: c.fill=PatternFill("solid",fgColor=GOLD)
wb.save(F)
print("R0187789 : mean %.4f  weighted %.4f  median %.2f"%(st.mean(a),sum(x*y for x,y in zip(a,asf))/sum(asf),st.median(a)))
print("R0188699 : mean %.4f (doc 7.78)  weighted %.4f (doc 7.38)  median %.2f"%(st.mean(b),sum(x*y for x,y in zip(b,bsf))/sum(bsf),st.median(b)))
print("R0200721 : sched mean %.4f (doc 9.30) median %.2f (doc 10.00) | eff mean %.4f (doc 8.41) median %.2f (doc 7.81) | sched weighted %.4f"%(
   st.mean(c5),st.median(c5),st.mean(e5),st.median(e5),sum(x*y for x,y in zip(c5,c5sf))/sum(c5sf)))
w=load_workbook(F); print("comp rows:",sum(1 for r in range(5,w['Lease Comparables'].max_row+1) if w['Lease Comparables'].cell(row=r,column=2).value not in (None,"")))
