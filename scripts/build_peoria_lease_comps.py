# -*- coding: utf-8 -*-
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import date

NAVY="FF1F3864"; BLUE="FF2E5C8A"; LTBLUE="FFD9E2F3"; GREEN="FFE2EFDA"; GOLD="FFFFF2CC"; GREY="FFF2F2F2"
W=Font(name="Arial",size=10,color="FFFFFFFF",bold=True)
TITLE=Font(name="Arial",size=12,color="FFFFFFFF",bold=True)
B=Font(name="Arial",size=10,bold=True)
N=Font(name="Arial",size=10)
I=Font(name="Arial",size=9,italic=True,color="FF595959")
thin=Side(style="thin",color="FFBFBFBF")
BOX=Border(left=thin,right=thin,top=thin,bottom=thin)
MONEY='"$"#,##0'; M2='"$"#,##0.00'; PCT='0.0%'; SF='#,##0" SF"'; NUM='#,##0'; DT='mm/dd/yyyy'

wb=Workbook()

def title(ws,row,ncols,text,sub=None):
    ws.merge_cells(start_row=row,start_column=1,end_row=row,end_column=ncols)
    c=ws.cell(row=row,column=1,value=text); c.font=TITLE; c.fill=PatternFill("solid",fgColor=NAVY)
    c.alignment=Alignment(horizontal="left",vertical="center"); ws.row_dimensions[row].height=22
    if sub:
        ws.merge_cells(start_row=row+1,start_column=1,end_row=row+1,end_column=ncols)
        c=ws.cell(row=row+1,column=1,value=sub); c.font=I
        return row+2
    return row+1

def hdr(ws,row,vals,fill=BLUE):
    for j,v in enumerate(vals,1):
        c=ws.cell(row=row,column=j,value=v); c.font=W; c.fill=PatternFill("solid",fgColor=fill)
        c.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True); c.border=BOX
    ws.row_dimensions[row].height=32
    return row+1

def widths(ws,ws_w):
    for i,w in enumerate(ws_w,1): ws.column_dimensions[get_column_letter(i)].width=w

# ------------------------------------------------------------------ COMP DATA
COMPS=[
 dict(no=1,name="Century Square Shopping Center",addr="10300-10498 E Colfax Ave",city="Aurora",zipc="80010",
   submkt="Central",signed=date(2024,2,7),comm=date(2024,3,8),dtype="Direct",newr="New Lease",sf=1285,pctbldg=.117,
   basis="Asking",rent=30.00,ask=30.00,start=None,svc="NNN",cam=7.00,camann=8995,annual=38550,term=None,exp=None,
   tenant=None,landlord="The Daniel Group",broker="Daniel Group Ltd The",
   bcls="C",built="1985/2008",bsf=10943,land=41818,landac=0.96,btype="Storefront",tenancy="Multi",
   pos="Outparcel",cond="Excellent",bldout="Shell Space",tom=20.0,vac=21.0,
   subvac=0.000,subrent=32.44,smvac=.041,smrent=25.83,mkvac=.038,mkrent=26.12,pg="13-14"),
 dict(no=2,name="Hoffman Heights Shopping Center",addr="710-750 Peoria St",city="Aurora",zipc="80011",
   submkt="Aurora",signed=date(2023,11,14),comm=date(2023,12,14),dtype="Sublet",newr="New Lease",sf=3750,pctbldg=.052,
   basis="Asking",rent=15.00,ask=15.00,start=None,svc="NNN",cam=5.95,camann=22313,annual=56250,term=None,exp=None,
   tenant=None,landlord="Goodman Realty Group",broker="Antonoff & Company / Fairbairn Commercial",
   bcls="B",built="1960/1995",bsf=71692,land=566280,landac=13.00,btype="Storefront",tenancy="Multi",
   pos=None,cond=None,bldout=None,tom=6.4,vac=7.4,
   subvac=.067,subrent=15.16,smvac=.028,smrent=19.56,mkvac=.041,mkrent=25.95,pg="15-16"),
 dict(no=3,name="Hoffman Heights Shopping Center",addr="710-750 Peoria St",city="Aurora",zipc="80011",
   submkt="Aurora",signed=date(2023,11,14),comm=date(2023,12,14),dtype="Sublet",newr="New Lease",sf=1100,pctbldg=.015,
   basis="Asking",rent=25.00,ask=25.00,start=None,svc="NNN",cam=5.95,camann=6545,annual=27500,term=None,exp=None,
   tenant=None,landlord="Goodman Realty Group",broker="Antonoff & Company / Fairbairn Commercial",
   bcls="B",built="1960/1995",bsf=71692,land=566280,landac=13.00,btype="Storefront",tenancy="Multi",
   pos="End Cap",cond=None,bldout=None,tom=6.1,vac=1.4,
   subvac=.067,subrent=15.16,smvac=.028,smrent=19.56,mkvac=.041,mkrent=25.95,pg="17-18"),
 dict(no=4,name="10690 Del Mar Pky",addr="10690 Del Mar Pky",city="Aurora",zipc="80010",
   submkt="Aurora",signed=date(2023,4,19),comm=date(2023,4,19),dtype="Direct",newr="New Lease",sf=4250,pctbldg=.113,
   basis="Starting",rent=15.00,ask=15.00,start=15.00,svc="NNN",cam=None,camann=None,annual=63750,term=10.0,exp=date(2033,4,19),
   tenant="Community Medical Services",landlord="Excell Fund Brokerage",broker="Excell Fund Brokerage (Cresa - tenant rep)",
   bcls="C",built="1960",bsf=37500,land=149628,landac=3.43,btype="Retail",tenancy="Multi",
   pos="In-Line",cond="Average",bldout="Full Standard Retail Build-Out",tom=3.3,vac=2.6,
   subvac=0.000,subrent=12.96,smvac=.035,smrent=18.93,mkvac=.039,mkrent=25.42,pg="19-20"),
 dict(no=5,name="Hoffman Heights Shopping Center",addr="710-750 Peoria St",city="Aurora",zipc="80011",
   submkt="Aurora",signed=date(2023,3,30),comm=date(2023,4,1),dtype="Sublet",newr="New Lease",sf=3750,pctbldg=.052,
   basis="Asking",rent=15.00,ask=15.00,start=None,svc="NNN",cam=5.42,camann=20325,annual=56250,term=None,exp=None,
   tenant=None,landlord="Goodman Realty Group",broker="Antonoff & Company / Fairbairn Commercial",
   bcls="B",built="1960/1995",bsf=71692,land=566280,landac=13.00,btype="Storefront",tenancy="Multi",
   pos=None,cond=None,bldout=None,tom=6.5,vac=3.0,
   subvac=.186,subrent=14.55,smvac=.041,smrent=18.99,mkvac=.037,mkrent=25.23,pg="21-22"),
 dict(no=6,name="Hoffman Heights Shopping Center",addr="710-750 Peoria St",city="Aurora",zipc="80011",
   submkt="Aurora",signed=date(2023,3,20),comm=date(2023,4,1),dtype="Sublet",newr="New Lease",sf=3840,pctbldg=.054,
   basis="Starting",rent=12.50,ask=15.00,start=12.50,svc="NNN",cam=None,camann=None,annual=48000,term=5.0,exp=date(2028,4,1),
   tenant="Comunidad Latino LLC",landlord="Goodman Realty Group",broker="Antonoff & Company / Fairbairn (HomeSmart - tenant rep)",
   bcls="B",built="1960/1995",bsf=71692,land=566280,landac=13.00,btype="Storefront",tenancy="Multi",
   pos=None,cond=None,bldout=None,tom=6.7,vac=None,
   subvac=.186,subrent=14.55,smvac=.041,smrent=18.99,mkvac=.037,mkrent=25.23,pg="23-24"),
 dict(no=7,name="Bank of the West",addr="10401 E Colfax Ave",city="Aurora",zipc="80010",
   submkt="Aurora",signed=date(2022,8,16),comm=date(2022,9,1),dtype="Direct",newr="New Lease",sf=1886,pctbldg=.186,
   basis="Asking",rent=24.00,ask=24.00,start=None,svc="MG",cam=None,camann=None,annual=45264,term=None,exp=None,
   tenant=None,landlord=None,broker="CBRE",
   bcls="B",built="1979/2007",bsf=10157,land=37462,landac=0.86,btype="Office",tenancy="Multi",
   pos=None,cond=None,bldout=None,tom=8.8,vac=9.3,
   subvac=0.000,subrent=24.27,smvac=.092,smrent=24.15,mkvac=.146,mkrent=28.99,pg="25"),
 dict(no=8,name="1408 N Del Mar Pky",addr="1408 N Del Mar Pky",city="Aurora",zipc="80010",
   submkt="Central",signed=date(2022,6,30),comm=date(2022,7,30),dtype="Direct",newr="New Lease",sf=1418,pctbldg=.269,
   basis="Asking",rent=16.00,ask=16.00,start=None,svc="NNN",cam=None,camann=None,annual=22688,term=None,exp=None,
   tenant=None,landlord="Maria Del Carmen Mora",broker="Capital Property Group",
   bcls="C",built="1973",bsf=5278,land=11761,landac=0.27,btype="Storefront",tenancy="Multi",
   pos="End Cap",cond="Excellent",bldout="Full Standard Retail Build-Out",tom=29.0,vac=30.0,
   subvac=.269,subrent=16.52,smvac=.035,smrent=24.15,mkvac=.043,mkrent=24.44,pg="26-27"),
]
SRC="Source: Grease Monkey International, LLC - 1521-1527 Peoria St - CO Property Tax 2025/2026 Analysis (Ryan, LLC / CoStar), PDF pp. 11-27."
LOV=date(2024,6,30)

# ============================================================ 1. SUBJECT & CONCLUSION
# Built last (then moved to the front) so cross-sheet stat refs are computed, not hardcoded.
def build_subject():
 ws=wb.create_sheet("Subject & Conclusion")
 widths(ws,[34,20,16,16,16,16,14,14,14,14])
 r=title(ws,1,10,"SUBJECT PROPERTY & MARKET RENT CONCLUSION",
         "Grease Monkey International, LLC  |  1521-1527 Peoria St, Aurora, CO 80010  |  Adams County  |  Assessment Year 2025/2026")
 r+=1
 r=hdr(ws,r,["Subject Property Facts","Value","","","","","","","",""])
 facts=[("Client / Owner","Grease Monkey International, LLC"),("Property Name","1521-1527 Peoria St"),
  ("Address","1521-1527 Peoria St, Aurora, CO 80010"),("County","Adams County"),
  ("Parcel Number","182335429010"),("Property Type","Retail - Service Station"),
  ("Net Leasable Area (SF)",4052),("Year of Construction",1982),
  ("Site Size (Acres)",0.500),("Site Size (SF)",21780),("Land-to-Building Ratio","5.38 : 1"),
  ("Prepared By","Declan Fassig, Ryan LLC (Agent)")]
 for k,v in facts:
     ws.cell(row=r,column=1,value=k).font=B
     c=ws.cell(row=r,column=2,value=v); c.font=N
     if k.endswith("(SF)") or k=="Year of Construction": c.number_format=NUM
     if k=="Year of Construction": c.number_format="0"
     if k=="Site Size (Acres)": c.number_format="0.000"
     ws.cell(row=r,column=1).border=BOX; ws.cell(row=r,column=2).border=BOX
     r+=1
 r+=1
 r=hdr(ws,r,["Assessment / Value Summary","2025/2026","2023/2024","","","","","","",""])
 av=[("Land Value",453024,383328),("Improvement Value",707317,791752),("Total Actual Value",1160341,1175080)]
 for k,a,b in av:
     ws.cell(row=r,column=1,value=k).font=B if k.startswith("Total") else N
     for j,v in ((2,a),(3,b)):
         c=ws.cell(row=r,column=j,value=v); c.number_format=MONEY; c.font=B if k.startswith("Total") else N; c.border=BOX
         if k.startswith("Total"): c.fill=PatternFill("solid",fgColor=GREEN)
     ws.cell(row=r,column=1).border=BOX
     if k.startswith("Total"): ws.cell(row=r,column=1).fill=PatternFill("solid",fgColor=GREEN)
     r+=1
 for k,a,b in [("Total Value PSF",286.36,290.00),("Land Value PSF",20.80,17.60)]:
     ws.cell(row=r,column=1,value=k).font=N; ws.cell(row=r,column=1).border=BOX
     for j,v in ((2,a),(3,b)):
         c=ws.cell(row=r,column=j,value=v); c.number_format=M2; c.font=N; c.border=BOX
     r+=1
 ws.cell(row=r,column=1,value="Value Change from Prior Year").font=N; ws.cell(row=r,column=1).border=BOX
 c=ws.cell(row=r,column=2,value=-0.0125); c.number_format=PCT; c.font=N; c.border=BOX
 r+=2

 r=hdr(ws,r,["Taxpayer Proforma Income Analysis (PDF p. 3)","Annual","Per SF","% PGI / % EGI","","","","","",""])
 pro=[("Potential Base Rent",81040,20.00,1.0000,"% PGI"),("Gross Potential Rent",81040,20.00,1.0000,"% PGI"),
  ("Less: Vacancy Loss",-4862,-1.20,0.0600,"% PGI"),("Effective Gross Income",76178,18.80,0.9400,"% PGI"),
  ("Less: Total Operating Expenses",-6094,-1.50,0.0800,"% EGI"),("NET OPERATING INCOME",70083,17.30,0.9200,"% EGI")]
 for k,a,psf,p,_ in pro:
     tot=k in("Gross Potential Rent","Effective Gross Income","NET OPERATING INCOME")
     f=B if tot else N
     ws.cell(row=r,column=1,value=k).font=f
     c=ws.cell(row=r,column=2,value=a); c.number_format=MONEY; c.font=f
     c2=ws.cell(row=r,column=3,value=psf); c2.number_format=M2; c2.font=f
     c3=ws.cell(row=r,column=4,value=p); c3.number_format=PCT; c3.font=f
     for j in range(1,5):
         ws.cell(row=r,column=j).border=BOX
         if tot: ws.cell(row=r,column=j).fill=PatternFill("solid",fgColor=GREEN)
     r+=1
 for k,v,fmt in [("Capitalization Rate",0.0750,PCT),("Indicated FMV",934445,MONEY),
                 ("Indicated FMV of RE (Rounded)",934400,MONEY),("Indicated FMV Per SF",230.60,M2)]:
     ws.cell(row=r,column=1,value=k).font=B; ws.cell(row=r,column=1).border=BOX
     c=ws.cell(row=r,column=2,value=v); c.number_format=fmt; c.font=B; c.border=BOX
     c.fill=PatternFill("solid",fgColor=GOLD)
     r+=1
 ws.cell(row=r,column=1,value="Note per source: market rent derived using the attached lease comps; cap rate derived from market reports (Denver community retail 7.50%).").font=I
 r+=2

 r=hdr(ws,r,["MARKET RENT CONCLUSION - Comparable Support","Rent $/SF (NNN)","Indication vs. Subject $20.00/SF","","","","","","",""])
 concl=[("Raw asking rent - 8 comps (average)",f"=Comparables!S{REF['Asking Rent $/SF']}",None),
        ("Raw asking rent - 8 comps (median)",f"=Comparables!T{REF['Asking Rent $/SF']}",None),
        ("Adjusted NNN market rent - average of 8 comps",f"='Adjustment Grid'!N{GREF['AVERAGE - adjusted NNN rent (all 8)']}",None),
        ("Adjusted NNN market rent - median of 8 comps",f"='Adjustment Grid'!N{GREF['MEDIAN - adjusted NNN rent (all 8)']}",None),
        ("Subject proforma base rent (as filed)",20.00,None)]
 for k,v,_ in concl:
     bold = "Adjusted" in k or "Subject" in k
     ws.cell(row=r,column=1,value=k).font=B if bold else N; ws.cell(row=r,column=1).border=BOX
     c=ws.cell(row=r,column=2,value=v); c.number_format=M2; c.font=B if bold else N; c.border=BOX
     if bold: c.fill=PatternFill("solid",fgColor=GOLD)
     d=ws.cell(row=r,column=3); d.border=BOX
     if "Subject" not in k:
         d.value=f"=IF(B{r}=0,\"\",B{r}/20-1)"; d.number_format='+0.0%;-0.0%;0.0%'; d.font=N
     r+=1
 r+=1
 ws.cell(row=r,column=1,value=SRC).font=I; r+=1
 ws.cell(row=r,column=1,value="Cells shown as formulas populate on open in Excel / Google Sheets.").font=I
 ws.freeze_panes="A4"

# ============================================================ 2. COMPARABLES
ws=wb.create_sheet("Comparables")
cols=["Comp #","Property Name","Address","City","ZIP","Submarket","Sign Date","Commencement","Deal Type",
 "Lease Type","Space Use","SF Leased","% of Building","Rent Basis","Asking Rent $/SF","Starting Rent $/SF",
 "Rent $/SF Used","Annual Rent $","Service Type","CAM $/SF","CAM Annual $","Gross-Equiv. $/SF","Term (Yrs)",
 "Expiration","Tenant","Landlord","Leasing Company","Bldg Class","Built / Renov.","Bldg Size (SF)",
 "Land Area (SF)","Land (AC)","Bldg Type","Tenancy","Space Position","Condition","Build-Out",
 "Months on Market","Months Vacant","Source Pages"]
widths(ws,[7,30,26,10,8,11,12,13,10,11,10,10,11,10,12,12,12,13,11,10,12,13,9,12,26,24,34,8,12,13,13,10,11,9,13,11,26,11,11,12])
r=title(ws,1,len(cols),"LEASE COMPARABLES - EXTRACTED DETAIL",
        "8 retail lease transactions, Aurora CO (CoStar via Ryan, LLC). Signed June 2022 - February 2024.")
r+=1
hr=r; r=hdr(ws,r,cols)
first=r
for cp in COMPS:
    ge = cp["rent"] + (cp["cam"] or 0) if cp["svc"]=="NNN" else cp["rent"]
    vals=[cp["no"],cp["name"],cp["addr"],cp["city"],cp["zipc"],cp["submkt"],cp["signed"],cp["comm"],cp["dtype"],
     cp["newr"],"Retail",cp["sf"],cp["pctbldg"],cp["basis"],cp["ask"],cp["start"],cp["rent"],cp["annual"],
     cp["svc"],cp["cam"],cp["camann"],ge,cp["term"],cp["exp"],cp["tenant"],cp["landlord"],cp["broker"],
     cp["bcls"],cp["built"],cp["bsf"],cp["land"],cp["landac"],cp["btype"],cp["tenancy"],cp["pos"],cp["cond"],
     cp["bldout"],cp["tom"],cp["vac"],cp["pg"]]
    for j,v in enumerate(vals,1):
        c=ws.cell(row=r,column=j,value=v); c.font=N; c.border=BOX
        c.alignment=Alignment(vertical="top",wrap_text=(j in(2,3,25,26,27,37)))
        if j in(7,8,24): c.number_format=DT
        if j in(13,): c.number_format=PCT
        if j in(15,16,17,20,22): c.number_format=M2
        if j in(18,21): c.number_format=MONEY
        if j in(12,30,31): c.number_format=NUM
        if j in(23,38,39): c.number_format='0.0'
        if j==32: c.number_format='0.00'
    if cp["no"]%2==0:
        for j in range(1,len(cols)+1): ws.cell(row=r,column=j).fill=PatternFill("solid",fgColor=GREY)
    r+=1
last=r-1
# stats block
r+=1
ws.cell(row=r,column=1,value="SUMMARY STATISTICS").font=B
r+=1
sr=r
ws.cell(row=r,column=16,value="Metric").font=B
for j,lab in enumerate(["Count","Low","Average","Median","High"]):
    c=ws.cell(row=r,column=17+j,value=lab); c.font=W; c.fill=PatternFill("solid",fgColor=BLUE)
    c.alignment=Alignment(horizontal="center"); c.border=BOX
c=ws.cell(row=r,column=16); c.font=W; c.fill=PatternFill("solid",fgColor=BLUE); c.border=BOX
r+=1
stat_rows=[("Rent $/SF used (all 8)","Q",M2),("Asking Rent $/SF","O",M2),("SF Leased","L",NUM),
           ("Gross-Equivalent $/SF","V",M2),("Months on Market","AL",'0.0')]
REF={}
for lab,col,fmt in stat_rows:
    REF[lab]=r
    ws.cell(row=r,column=16,value=lab).font=B; ws.cell(row=r,column=16).border=BOX
    rng=f"{col}{first}:{col}{last}"
    for j,f in enumerate([f"=COUNT({rng})",f"=MIN({rng})",f"=AVERAGE({rng})",f"=MEDIAN({rng})",f"=MAX({rng})"]):
        c=ws.cell(row=r,column=17+j,value=f); c.font=N; c.border=BOX
        c.number_format=NUM if j==0 else fmt
        c.fill=PatternFill("solid",fgColor=GREEN)
    r+=1
r+=1
ws.cell(row=r,column=16,value="CoStar-published summary statistics (PDF p. 11), for cross-check:").font=B; r+=1
for lab,vals in [("Asking Rent per SF",[8,15.00,19.38,15.50,30.00]),("Starting Rent per SF",[2,12.50,13.75,13.75,15.00]),
                 ("Asking Rent Discount",[2,0.000,0.083,0.083,0.167]),("Lease Size (SF)",[8,1100,2660,2818,4250]),
                 ("Term (Years)",[2,5.0,7.5,7.5,10.0]),("Months on Market",[8,3,11,6,29])]:
    ws.cell(row=r,column=16,value=lab).font=N; ws.cell(row=r,column=16).border=BOX
    for j,v in enumerate(vals):
        c=ws.cell(row=r,column=17+j,value=v); c.font=N; c.border=BOX
        c.number_format=NUM if j==0 else (PCT if "Discount" in lab else (NUM if "Size" in lab else ('0.0' if ("Term" in lab or "Months" in lab) else M2)))
    r+=1
r+=1
ws.cell(row=r,column=1,value="Note: Effective rent, improvement allowance and free rent were not reported for any of the 8 transactions.").font=I; r+=1
ws.cell(row=r,column=1,value=SRC).font=I
ws.freeze_panes=f"B{first}"
ws.auto_filter.ref=f"A{hr}:{get_column_letter(len(cols))}{last}"

# ============================================================ 3. ADJUSTMENT GRID
ws=wb.create_sheet("Adjustment Grid")
widths(ws,[7,32,11,10,10,13,13,13,13,13,13,13,13,14,40])
r=title(ws,1,15,"LEASE COMPARABLE ADJUSTMENT GRID -> INDICATED MARKET RENT",
        "Level-of-value date 06/30/2024 (Colorado 2025/2026 reassessment). All indications on a triple-net (NNN) basis.")
r+=1
# assumptions
ws.cell(row=r,column=1,value="ADJUSTMENT ASSUMPTIONS (editable - grid recalculates)").font=B
ASSUM=r+1
aslist=[("Adj 1 - Asking-to-starting rent discount",0.083,PCT,"CoStar-reported average asking-rent discount for this comp set (PDF p. 11); applied to asking-rent-only comps."),
 ("Adj 2 - MG-to-NNN expense load ($/SF)",6.08,M2,"Average reported CAM of the four comps disclosing it ($7.00, $5.95, $5.95, $5.42). Deducted from the one Modified Gross comp."),
 ("Adj 3 - Market conditions, annual rent growth",0.035,PCT,"Submarket asking-rent YOY growth reported across the comp set (3.2%-6.6%); 3.5% applied straight-line to 06/30/2024."),
 ("Adj 4 - Size premium, spaces under 2,000 SF",-0.05,PCT,"Small-suite $/SF premium removed; subject NLA is 4,052 SF."),
 ("Sublet transaction adjustment",0.000,PCT,"Set to 0%; sublets are flagged as secondary reliability rather than quantitatively adjusted.")]
rr=ASSUM
for k,v,fmt,note in aslist:
    ws.cell(row=rr,column=1,value=k).font=N
    ws.merge_cells(start_row=rr,start_column=1,end_row=rr,end_column=2)
    c=ws.cell(row=rr,column=3,value=v); c.number_format=fmt; c.font=B; c.fill=PatternFill("solid",fgColor=GOLD); c.border=BOX
    ws.cell(row=rr,column=4,value=note).font=I
    rr+=1
r=rr+1
gcols=["Comp #","Property / Address","SF Leased","Sign Date","Rent Basis","Quoted Rent $/SF","Service Type",
 "Adj 1: Rent Basis","Adj 2: Expense Basis to NNN","Adj 3: Market Conditions","Adj 4: Size","Net $ Adjustment",
 "Total Adj %","Adjusted NNN Rent $/SF","Reliability / Comment"]
hr=r; r=hdr(ws,r,gcols)
gfirst=r
for cp in COMPS:
    a1=ASSUM; a2=ASSUM+1; a3=ASSUM+2; a4=ASSUM+3
    yrs=(LOV-cp["signed"]).days/365.25
    rel = "Sublet - secondary support" if cp["dtype"]=="Sublet" else "Direct lease - primary support"
    extra=[]
    if cp["basis"]=="Starting": extra.append("actual contract rent")
    if cp["svc"]=="MG": extra.append("MG converted to NNN")
    if cp["sf"]<2000: extra.append("small suite")
    if cp["btype"]=="Office": extra.append("office-type bldg")
    if extra: rel += " (" + "; ".join(extra) + ")"
    vals=[cp["no"],f'{cp["name"]} - {cp["addr"]}',cp["sf"],cp["signed"],cp["basis"],cp["rent"],cp["svc"],
      (f"=-F{r}*$C${a1}" if cp["basis"]=="Asking" else 0),
      (f"=-$C${a2}" if cp["svc"]=="MG" else 0),
      f"=(F{r}+H{r}+I{r})*$C${a3}*{yrs:.3f}",
      (f"=(F{r}+H{r}+I{r}+J{r})*$C${a4}" if cp["sf"]<2000 else 0),
      f"=SUM(H{r}:K{r})", f"=L{r}/F{r}", f"=F{r}+L{r}", rel]
    for j,v in enumerate(vals,1):
        c=ws.cell(row=r,column=j,value=v); c.font=N; c.border=BOX
        c.alignment=Alignment(vertical="center",wrap_text=(j in(2,15)))
        if j==3: c.number_format=NUM
        if j==4: c.number_format=DT
        if j in(6,8,9,10,11,12,14): c.number_format='"$"#,##0.00;[Red]("$"#,##0.00)'
        if j==13: c.number_format='+0.0%;-0.0%;0.0%'
        if j==14: c.font=B; c.fill=PatternFill("solid",fgColor=LTBLUE)
    if cp["no"]%2==0:
        for j in [1,2,3,4,5,6,7,8,9,10,11,12,13,15]: ws.cell(row=r,column=j).fill=PatternFill("solid",fgColor=GREY)
    r+=1
glast=r-1
r+=1
stat_start=r
GREF={}
for lab,fn in [("Count of comparables",f"=COUNT(N{gfirst}:N{glast})"),
  ("Low - adjusted NNN rent",f"=MIN(N{gfirst}:N{glast})"),
  ("AVERAGE - adjusted NNN rent (all 8)",f"=AVERAGE(N{gfirst}:N{glast})"),
  ("MEDIAN - adjusted NNN rent (all 8)",f"=MEDIAN(N{gfirst}:N{glast})"),
  ("High - adjusted NNN rent",f"=MAX(N{gfirst}:N{glast})"),
  ("Average - DIRECT leases only (comps 1, 4, 7, 8)",f"=AVERAGE(N{gfirst},N{gfirst+3},N{gfirst+6},N{gfirst+7})"),
  ("Average - SUBLET leases only (comps 2, 3, 5, 6)",f"=AVERAGE(N{gfirst+1},N{gfirst+2},N{gfirst+4},N{gfirst+5})"),
  ("Average - comps 3,000+ SF (closest to subject 4,052 SF)",f"=AVERAGE(N{gfirst+1},N{gfirst+3},N{gfirst+4},N{gfirst+5})")]:
    GREF[lab]=r
    ws.cell(row=r,column=12,value=lab).font=B
    ws.merge_cells(start_row=r,start_column=12,end_row=r,end_column=13)
    c=ws.cell(row=r,column=14,value=fn); c.font=B; c.border=BOX
    c.number_format=NUM if "Count" in lab else M2
    c.fill=PatternFill("solid",fgColor=GREEN)
    ws.cell(row=r,column=12).border=BOX
    r+=1
r+=1
ws.cell(row=r,column=12,value="Subject proforma base rent as filed").font=B
ws.merge_cells(start_row=r,start_column=12,end_row=r,end_column=13)
c=ws.cell(row=r,column=14,value=20.00); c.number_format=M2; c.font=B; c.fill=PatternFill("solid",fgColor=GOLD); c.border=BOX
subj_row=r; r+=1
ws.cell(row=r,column=12,value="Filed rent vs. adjusted average").font=B
ws.merge_cells(start_row=r,start_column=12,end_row=r,end_column=13)
c=ws.cell(row=r,column=14,value=f"=N{subj_row}/N{stat_start+2}-1"); c.number_format='+0.0%;-0.0%;0.0%'; c.font=B; c.border=BOX
r+=2
for line in [
 "Adjustment method:",
 "  Adj 1 - Rent Basis: asking rents are reduced by the comp set's own reported asking-rent discount to approximate contract (starting) rent. Comps 4 and 6 report actual starting rent and take no adjustment.",
 "  Adj 2 - Expense Basis: comp 7 is quoted Modified Gross; the average reported CAM load is deducted to state it on the same NNN basis as the subject and the other seven comps.",
 "  Adj 3 - Market Conditions: straight-line rent growth from each sign date to the 06/30/2024 level-of-value date.",
 "  Adj 4 - Size: comps under 2,000 SF (1, 3, 7, 8) carry a small-suite premium that the subject's 4,052 SF does not command.",
 "  Not adjusted: location (all comps are Aurora / E Colfax - Peoria corridor, same trade area as the subject), space use (all retail), and floor (all ground floor).",
 "",
 "Reliability note: four of the eight transactions are sublets at a single property (Hoffman Heights, 710-750 Peoria St) and are asking rather than contract rents; they are directionally useful but carry less weight than the direct leases.",
 "The subject is a freestanding service-station / quick-lube improvement, while every comparable is in-line or outparcel multi-tenant retail. No like-use (automotive service) comparable is included in the source package.",
 SRC]:
    c=ws.cell(row=r,column=1,value=line); c.font=I if line else N
    r+=1
ws.freeze_panes=f"C{gfirst}"

# ============================================================ 4. MARKET CONDITIONS
ws=wb.create_sheet("Market Conditions")
mc=["Comp #","Property / Address","Period","Subject Vacancy","Subject YOY","Submarket Vacancy",
 "Submarket Rent $/SF","Submarket Rent YOY","Market Vacancy","Market Rent $/SF","Market Rent YOY",
 "12-Mo Leased (SF)","Months on Market","12-Mo Sales Volume","Sale Price $/SF"]
widths(ws,[7,44,10,13,11,13,14,14,12,13,12,15,13,16,13])
r=title(ws,1,len(mc),"MARKET CONDITIONS REPORTED WITH EACH COMPARABLE",
        "CoStar submarket and market statistics as of each comparable's transaction quarter.")
r+=1
r=hdr(ws,r,mc)
MKT=[(1,"2024 Q1",0.000,-0.117,.041,25.83,.032,.038,26.12,.035,429250,7.2,85490000,310),
     (2,"2023 Q4",.067,-0.080,.028,19.56,.042,.041,25.95,.039,319861,12.2,55740000,275),
     (3,"2023 Q4",.067,-0.080,.028,19.56,.042,.041,25.95,.039,319861,12.2,55740000,275),
     (4,"2023 Q2",0.000,0.000,.035,18.93,.043,.039,25.42,.040,274237,10.4,35960000,407),
     (5,"2023 Q1",.186,0.186,.041,18.99,.044,.037,25.23,.043,243351,7.4,56910000,249),
     (6,"2023 Q1",.186,0.186,.041,18.99,.044,.037,25.23,.043,243351,7.4,54950000,252),
     (7,"2022 Q3",0.000,0.000,.092,24.15,.022,.146,28.99,.013,472940,15.6,51710000,99),
     (8,"2022 Q2",.269,0.000,.035,24.15,.035,.043,24.44,.040,440639,9.2,270570000,283)]
for row in MKT:
    cp=COMPS[row[0]-1]
    vals=[row[0],f'{cp["name"]} - {cp["addr"]}',row[1],row[2],row[3],row[4],row[5],row[6],row[7],row[8],row[9],row[10],row[11],row[12],row[13]]
    for j,v in enumerate(vals,1):
        c=ws.cell(row=r,column=j,value=v); c.font=N; c.border=BOX
        c.alignment=Alignment(vertical="center",wrap_text=(j==2))
        if j in(4,6,9): c.number_format=PCT
        if j in(5,8,11): c.number_format='+0.0%;-0.0%;0.0%'
        if j in(7,10): c.number_format=M2
        if j==12: c.number_format=NUM
        if j==13: c.number_format='0.0'
        if j==14: c.number_format=MONEY
        if j==15: c.number_format='"$"#,##0"/SF"'
    if row[0]%2==0:
        for j in range(1,len(mc)+1): ws.cell(row=r,column=j).fill=PatternFill("solid",fgColor=GREY)
    r+=1
r+=2
ws.cell(row=r,column=1,value="Denver retail capitalization rates (PDF p. 8)").font=B; r+=1
r=hdr(ws,r,["Category","Denver, CO","West Region","National","","","","","","","","","","",""])
for k,a,b,c_ in [("Community Retail",0.0750,0.0655,0.0650),("Neighborhood Retail",0.0659,0.0725,0.0726)]:
    ws.cell(row=r,column=1,value=k).font=N; ws.cell(row=r,column=1).border=BOX
    for j,v in ((2,a),(3,b),(4,c_)):
        cc=ws.cell(row=r,column=j,value=v); cc.number_format='0.00%'; cc.font=N; cc.border=BOX
    r+=1
ws.cell(row=r,column=1,value="The 7.50% rate applied in the taxpayer's proforma matches the Denver community-retail indication.").font=I
r+=2
ws.cell(row=r,column=1,value=SRC).font=I
ws.freeze_panes="C4"

# ============================================================ 5. SOURCE NOTES
ws=wb.create_sheet("Source Notes")
widths(ws,[26,22,90])
r=title(ws,1,3,"SOURCE INDEX, METHOD & DATA GAPS")
r+=1
r=hdr(ws,r,["Item","Reference","Detail"])
notes=[("Source document","PDF, 27 pages","Grease_Monkey_International_LLC_CO_PT_20252026_1521_Peoria_St_Analysis_20260805.pdf"),
 ("Subject / client","p. 1-3","Grease Monkey International, LLC - 1521-1527 Peoria St, Aurora CO 80010, Adams County, parcel 182335429010"),
 ("Preparer","p. 1","Declan Fassig, Ryan LLC, acting as agent - 720.303.5198 - declan.fassig@ryan.com"),
 ("Comparable data","p. 11-27","CoStar Group lease comps, licensed to Ryan, LLC (license 1702672), run 8/5/2026"),
 ("Comps extracted","8 of 8","All lease transactions in the source package were captured; none were omitted."),
 ("Comp map","p. 11","Map graphic only - no distance-to-subject figures published; not extracted."),
 ("Excluded from this workbook","p. 4-10","Market overview, SitusAMC narrative, cap-rate charts (p. 8 summarized on Market Conditions tab) and blank pages 9-10."),
 ("Level of value date","06/30/2024","Colorado 2025/2026 reassessment cycle appraisal date; used for the market-conditions adjustment."),
 ("DATA GAP - effective rent","p. 11","Effective rent per SF not reported for any of the 8 deals (0 of 8)."),
 ("DATA GAP - concessions","p. 11","Improvement allowance and months of free rent not reported for any deal (0 of 8)."),
 ("DATA GAP - lease term","p. 11","Term reported for only 2 of 8 deals (comp 4 = 10 yrs, comp 6 = 5 yrs)."),
 ("DATA GAP - starting rent","p. 11","Contract/starting rent reported for only 2 of 8 deals; the other 6 are asking rents."),
 ("DATA GAP - CAM","p. 13-27","CAM disclosed for 4 of 8 deals ($7.00, $5.95, $5.95, $5.42 per SF)."),
 ("DATA GAP - tenant name","p. 13-27","Tenant identified for only 2 of 8 deals (Community Medical Services; Comunidad Latino LLC)."),
 ("CONCENTRATION","comps 2,3,5,6","Half the comp set is sublease activity at one property, Hoffman Heights Shopping Center, 710-750 Peoria St."),
 ("USE MISMATCH","all comps","Subject is a freestanding service station / quick-lube; all 8 comps are in-line or outparcel multi-tenant retail. No automotive-service comp is in the package."),
 ("Formulas","all tabs","Summary, statistic and adjustment cells are live formulas; they populate on open in Excel or Google Sheets.")]
for a,b_,c_ in notes:
    gap = a.startswith("DATA GAP") or a in("CONCENTRATION","USE MISMATCH")
    for j,v in enumerate([a,b_,c_],1):
        cc=ws.cell(row=r,column=j,value=v); cc.font=B if (gap and j==1) else N; cc.border=BOX
        cc.alignment=Alignment(vertical="top",wrap_text=True)
        if gap: cc.fill=PatternFill("solid",fgColor=GOLD)
    r+=1
ws.freeze_panes="A4"

build_subject()
del wb["Sheet"]   # unused default sheet
wb.move_sheet("Subject & Conclusion", offset=-(len(wb.sheetnames)-1))
assert wb.sheetnames==["Subject & Conclusion","Comparables","Adjustment Grid","Market Conditions","Source Notes"], wb.sheetnames

out="/tmp/claude-0/-home-user-Expense-Lease-Extractor-/2b2efe21-3929-51eb-ba00-13704fd989ef/scratchpad/1521-1527_Peoria_St_Lease_Comparables_Analysis.xlsx"
wb.save(out)
print("saved",out)
