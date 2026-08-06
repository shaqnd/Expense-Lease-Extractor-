# -*- coding: utf-8 -*-
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
from copy import copy
P="/home/user/Expense-Lease-Extractor-/Master_Property_Database.xlsx"
wb=load_workbook(P); NS="(not stated in document)"
MON='"$"#,##0.00'; MON0='"$"#,##0'; PSF='"$"#,##0.00'
def sf_(ws,r,nc): return [copy(ws.cell(row=r,column=c)._style) for c in range(1,nc+1)]
def put(ws,row,vals,st,nf=None):
    for j,v in enumerate(vals,1):
        c=ws.cell(row=row,column=j); c.value=v
        if j-1<len(st): c._style=copy(st[j-1])
        if nf and j in nf: c.number_format=nf[j]
        c.alignment=Alignment(vertical="top",wrap_text=True)
def snap(ws,frm,nc):
    return [[(ws.cell(row=r,column=c).value,copy(ws.cell(row=r,column=c)._style),ws.cell(row=r,column=c).number_format)
             for c in range(1,nc+1)] for r in range(frm,ws.max_row+1)]
def rest(ws,at,buf):
    for i,rd in enumerate(buf):
        for j,(v,s,f) in enumerate(rd,1):
            c=ws.cell(row=at+i,column=j); c.value=v; c._style=copy(s); c.number_format=f
def ins(ws,at,rows,src,nc,nf=None):
    n=len(rows); moved=[m for m in list(ws.merged_cells.ranges) if m.min_row>=at]
    for m in moved: ws.unmerge_cells(str(m))
    buf=snap(ws,at,nc); st=sf_(ws,src,nc)
    for i,r in enumerate(rows): put(ws,at+i,r,st,nf)
    rest(ws,at+n,buf)
    for m in moved: ws.merge_cells(start_row=m.min_row+n,start_column=m.min_col,end_row=m.max_row+n,end_column=m.max_col)

# SOURCE DOCS
ws=wb["Source Documents"]
last=max(r for r in range(6,ws.max_row+1) if isinstance(ws.cell(row=r,column=1).value,(int,float)))
ins(ws,last+1,[
 [25,"R0187787.pdf","Assessor protest - income approach, 2 rent rolls, 2yr expenses, 6 lease comps, sale comps, LOA","5690 E. 56th Avenue, Commerce City (KEW Realty)","Tax Yr 2025 (val 6/30/2024)","Sterling Property Tax Specialists (Paul Leonard)"],
 [26,"R0157664.pdf","Protest inquiry + LOA + 4-year income analysis + 2024/2023 and 2022/2021 income statements + May 2024 rent roll","6045 E 76th Avenue, Commerce City (4711 West Tennessee Avenue LLC)","Tax Yr 2025 (2021-2024 actuals)","R.H. Jacobson & Company"],
 [27,"R0164588.pdf","Tax appeal - market proforma + 5 lease comps + vacant rent roll + income statements","9410 Heinz Way, Commerce City (ASB Real Estate)","Tax Yr 2025 (val 6/30/2024)","Ryan LLC (Ethan Horn) / CoStar"],
 [28,"R0164287.pdf","Owner-filed CBOE appeal + 2024 county I&E survey + 2010-2025 rent schedule + income work sheet","2850 Walden Street, Aurora (EDC Holdings LLC)","Tax Yr 2025 (rent history 2010-2025)","EDC Holdings LLC (Edward D Chandler, owner)"],
 [29,"R0180894.pdf","Tax appeal - proforma income analysis + 3-year actual income trends + LOA","Park 70, 1910 N Gun Club Rd, Aurora (Alpha Industrial Properties)","Assess. Yr 2025 (2022-2024 actuals)","Ryan LLC (Beth Diehl)"],
],6,6)

# PROPERTY MASTER
ws=wb["Property Master"]; at=ws.max_row+1; st=sf_(ws,5,25)
pm=[
 ["R0187787","R0187787","R0187787","5690 E. 56th Avenue","5690 E. 56th Avenue","Commerce City","CO",NS,"Adams",
  "KEW Realty Corporation (David Spira, President)","KEW Realty Corporation","Industrial / Flex (multi-tenant)","Multi (4 units)",2.14,None,None,35000,2016,None,
  "100% (6/30/2024) - stabilized",5266489,"2025",3930000,"Sterling Property Tax Specialists / Goldstein Law Firm (agents)","R0187787.pdf"],
 ["R0157664","R0157664","0172132216014","Alpine Park II - 6045 E 76th Avenue","6045 E 76th Avenue","Commerce City","CO",NS,"Adams",
  "4711 West Tennessee Avenue LLC (Scott Alan McCormick, Manager)","4711 West Tennessee Avenue LLC","Industrial - Warehouse (multi-bay)","Multi (11 bays)",None,66700,None,21000,1988,None,
  "71.4% - 6,000 SF vacant of 21,000 SF (May 2024 rent roll)",3423218,"2025",2200000,"R.H. Jacobson & Company (agent)","R0157664.pdf"],
 ["R0164588","R0164588","R0164588","9410 Heinz Way","9410 Heinz Way","Commerce City","CO",NS,"Adams",
  "ASB Real Estate","9410 Heinz Owner LLC","Industrial - Distribution","Single tenant (vacant)",18.29,None,None,140630,2005,None,
  "0% - 100% vacant at 6/30/2024; Home Depot lease expired Nov 2022",17644243,"2025",8539000,"Allegiance / ASB Allegiance Fund","R0164588.pdf"],
 ["R0164287","R0164287","0182128401002","2850 Walden Street","2850 Walden Street","Aurora","CO","80011","Adams",
  "EDC Holdings LLC (Edward D Chandler, member)","EDC Holdings LLC","Industrial - Warehouse / Distribution (with admin offices)","Single tenant",None,None,None,27819,None,None,
  "100% - single tenant, 0% vacant",4531688,"2025",2835000,"Owner-managed","R0164287.pdf"],
 ["R0180894","R0180894","R0180894","Park 70 - 1910 N Gun Club Rd","1910 N Gun Club Rd","Aurora","CO","80019","Adams",
  "Alpha Industrial Properties (client); Cherry Owner II LLC (owner of record)","Cherry Owner II LLC","Industrial - Warehouse (Class A)","Not itemized",10.290,448232,None,163790,2019,"2.74 : 1",
  "100% (6/30 occupancy, all three actual years)",24584873,"2025",13214100,"Alpha Industrial Properties","R0180894.pdf"],
]
for i,r in enumerate(pm): put(ws,at+i,r,st,{14:'0.000',15:'#,##0',16:'#,##0',17:'#,##0',18:'0',21:MON0,23:MON0})

# TENANTS & LEASES
ws=wb["Tenants & Leases"]; at=ws.max_row+1; st=sf_(ws,6,13)
tl=[
 ["5690 E. 56th Avenue (R0187787)","5690A","NOVITECH INC","Actual - Modified Gross Industrial",10003.88,120046.56,12.00,"5/1/2017","7/31/2027","123-month term",None,None,
  "10,000 SF. Tenancy Schedule as of 6/30/2024. Annual recoveries $3.31 PSF; no misc income."],
 ["5690 E. 56th Avenue (R0187787)","5690C","CAMSO USA INC","Actual - Modified Gross Industrial",10000.00,120000.00,12.00,"8/1/2018","9/30/2028","122-month term",None,None,
  "10,000 SF. Annual recoveries $2.59 PSF."],
 ["5690 E. 56th Avenue (R0187787)","5690E","TECHNEAUX TECHNOLOGY SERVICES LLC","Actual - NNN Industrial",0.00,0.00,0.00,"5/1/2024","7/31/2027","39-month term",None,None,
  "5,000 SF. Rent roll shows $0.00 monthly and annual rent at 6/30/2024 - the lease had commenced two months earlier and was in a rent-abatement period. Annual recoveries $4.95 PSF. The letter cites this as the base-period lease that is 'in-line with market' while calling the three older leases above market."],
 ["5690 E. 56th Avenue (R0187787)","5690F","REBOUND TECHNOLOGIES INC","Actual - Modified Gross Industrial",9791.67,117500.04,11.75,"8/2/2018","12/31/2026","101-month term",None,None,
  "10,000 SF. Annual recoveries $2.59 PSF."],
 ["5690 E. 56th Avenue (R0187787)","TOTAL","4 units - 100% occupied","Actual - mixed MG / NNN",29795.55,357546.60,10.22,"-","-","-",None,None,
  "35,000 SF occupied, 0 SF vacant, 100.00%. Totals are the sum of the four scheduled rents; the Techneaux unit contributes $0 while abated, so PSF is depressed."],
 ["Alpine Park II (R0157664)","1, 2","Daisy Window Tint","Actual - 2 unit bay",3650.00,43800.00,14.60,"6/1/2020","Month-to-month","MTM",None,None,"3,000 SF. May 2024 rent roll."],
 ["Alpine Park II (R0157664)","3","ReTyres","Actual - 1 unit bay",1650.00,19800.00,13.20,"12/1/2016","Month-to-month","MTM",None,None,"1,500 SF."],
 ["Alpine Park II (R0157664)","4, 5","TAC Business dba Miracle Service","Actual - 2 unit bay",3700.00,44400.00,14.80,"7/1/2015","Month-to-month","MTM",None,None,"3,000 SF."],
 ["Alpine Park II (R0157664)","7","Drywall, Dorrance","Actual - 1 unit bay",1500.00,18000.00,12.00,"12/1/2016","Month-to-month","MTM",None,None,"1,500 SF."],
 ["Alpine Park II (R0157664)","9","Banchy Welding LLC","Actual - bay",2375.00,28500.00,19.00,"12/26/2022","12/31/2025","-",None,None,"1,500 SF. Labelled '2 unit Bay' on the rent roll but sized 1,500 SF - reproduced as printed."],
 ["Alpine Park II (R0157664)","12","Dave Mueller","Actual - 1 unit bay",1500.00,18000.00,12.00,"4/1/2017","Month-to-month","MTM",None,None,"1,500 SF."],
 ["Alpine Park II (R0157664)","13, 14","Karl M. Espinoza","Actual - 2 unit bay",4150.00,49800.00,16.60,"3/15/2023","3/31/2026","-",None,None,"3,000 SF. $150 increase scheduled 4/1/25."],
 ["Alpine Park II (R0157664)","6, 8, 10, 11","VACANT (4 bays)","-",None,None,None,"-","-","-",None,None,"6,000 SF vacant = 28.5% of the 21,000 SF building, as noted on the rent roll."],
 ["Alpine Park II (R0157664)","TOTAL","7 occupied bays of 11","Actual",18525.00,222300.00,14.82,"-","-","-",None,None,
  "Printed rent-roll totals: $18,525 scheduled monthly charges, $222,300 annual. PSF is on the 15,000 SF occupied, not the 21,000 SF building ($10.59 PSF building-wide)."],
 ["9410 Heinz Way (R0164588)","0100 (Bldg 01115A)","VACANT","-",0.00,0.00,0.00,"-","-","-",None,None,
  "140,630 SF, 100% vacant on the 12/31/2024 rent roll and at the 6/30/2024 date of value. Home Depot's lease expired November 2022; the appeal states the space has been listed for lease ~4 years."],
 ["2850 Walden Street (R0164287)","Whole","(single tenant - not named in the filing)","Actual - owner pays property tax only",17549.08,210589.00,7.57,"-","3 years remaining from 2024","20-year lease term",None,None,
  "27,819 SF. 2025 figures from the owner's rent schedule. The county survey ticks 'Gross' but the comments state the tenant pays every expense except property tax, so it operates as near-NNN. Rent history 2010-2025: $4.86 PSF (2010) rising to $7.57 (2025); 2023 $7.07 / $196,573; 2024 $7.35 / $204,455. The survey reports the building as 27,800 SF against 27,819 SF on the rent schedule and work sheet."],
]
for i,r in enumerate(tl): put(ws,at+i,r,st,{5:MON,6:MON,7:PSF,11:MON,12:MON})

# ASSESSMENTS
ws=wb["Assessments & Valuation"]
hdrB=[r for r in range(1,ws.max_row+1) if str(ws.cell(row=r,column=1).value or "").startswith("B. Income-Approach")][0]
lastB=max(r for r in range(hdrB+2,ws.max_row+1) if ws.cell(row=r,column=2).value is not None)
ins(ws,lastB+1,[
 ["5690 E. 56th Avenue - actual 2023 (R0187787)",445925,-35674,410251,-65287,344964,None,None],
 ["5690 E. 56th Avenue - actual 2024 (R0187787)",547382,-43791,503591,-86600,416991,None,None],
 ["5690 E. 56th Avenue - concluded, avg NOI (R0187787)",None,None,None,None,380978,0.0970,3930000],
 ["9410 Heinz Way - market proforma (R0164588)",773465,-77347,696119,-55689,640429,0.0750,8539000],
 ["Park 70 - market proforma (R0180894)",982740,-49137,933603,-74688,858915,0.0650,13214100],
 ["Park 70 - actual YE2024 (R0180894)",2146279,None,2146279,-1307976,838303,None,None],
 ["Park 70 - actual YE2023 (R0180894)",2148674,None,2148674,-1297289,851385,None,None],
 ["Park 70 - actual YE2022 (R0180894)",1635142,None,1635142,-765306,869836,None,None],
 ["Alpine Park II - stabilized, taxes in cap rate (R0157664)",260000,None,260000,-42000,218000,0.0980,2224490],
 ["Alpine Park II - stabilized, taxes in expenses (R0157664)",260000,None,260000,-107000,153000,0.0700,2185714],
 ["2850 Walden St - owner work sheet at requested value (R0164287)",210590,None,210590,-103260,107330,0.03785,2835663],
 ["2850 Walden St - owner work sheet at assessed value (R0164287)",210590,None,210590,-158922,51668,0.01141,4531688],
],hdrB+2,8,{2:MON0,3:MON0,4:MON0,5:MON0,6:MON0,7:'0.000%',8:MON0})
lastA=max(r for r in range(6,hdrB) if ws.cell(row=r,column=2).value is not None)
ins(ws,lastA+1,[
 ["5690 E. 56th Avenue (R0187787)","2025",5266489,150.47,3930000,112.29,None,"R0187787"],
 ["Alpine Park II (R0157664)","2025",3423218,163.01,2200000,104.76,None,"R0157664"],
 ["9410 Heinz Way (R0164588)","2025",17644243,125.47,8539000,60.72,12800000,"R0164588"],
 ["2850 Walden Street (R0164287)","2025",4531688,162.90,2835000,101.91,2834700,"R0164287"],
 ["Park 70 (R0180894)","2025",24584873,150.10,13214100,80.68,15521000,"R0180894"],
],6,8,{3:MON0,4:PSF,5:MON0,6:PSF,7:MON0})

# I&E SUMMARY
ws=wb["Income-Expense Summary"]
at=max(r for r in range(5,ws.max_row+1) if ws.cell(row=r,column=1).value and ws.cell(row=r,column=5).value)+1
ies=[
 ["5690 E. 56th Avenue - CY2023",445925,169139,"Owner-reported gross income and expenses. Expenses INCLUDE real estate taxes ($116,160); the petitioner's NOI of $344,964 strips taxes out, adds a 3% reserve and applies 8% vacancy. R0187787"],
 ["5690 E. 56th Avenue - CY2024",547382,194687,"Includes real estate taxes ($123,195); petitioner's NOI $416,991. Two-year average NOI $380,978. R0187787"],
 ["Alpine Park II - CY2024",229531,107553.30,"Accrual. TRUE NOI $121,977.70 - no debt service or depreciation. Expenses include real property taxes $69,401.60. R0157664"],
 ["Alpine Park II - CY2023",256200,91938.18,"Accrual. TRUE NOI $164,261.82. Real property taxes $59,985.48. R0157664"],
 ["Alpine Park II - CY2022",258992.67,96393.71,"Accrual. NOI $162,598.96; net income after $5,791.72 of capital improvements = $156,807.24. R0157664"],
 ["Alpine Park II - CY2021",289136.52,74778.87,"Accrual. NOI $214,357.65; net income after $2,642.04 of capital improvements = $211,715.61. R0157664"],
 ["Park 70 - YE2024",2146279,1307976,"Actual. TRUE NOI $838,303 ($5.13 PSF). Expenses include property taxes $912,402 and $1,118 of non-recoverable expense. 100% occupied. R0180894"],
 ["Park 70 - YE2023",2148674,1297289,"Actual. NOI $851,385 ($5.21 PSF). Property taxes $923,586; non-recoverable $17,221. R0180894"],
 ["Park 70 - YE2022",1635142,765306,"Actual. NOI $869,836 ($5.32 PSF). Property taxes $533,285; non-recoverable $8,751. R0180894"],
 ["2850 Walden Street - 2025 (owner schedule)",210589,158923,"Owner's rent schedule. The only expense the owner bears is property tax; the tenant pays everything else. Net $51,666 - a 1.14% return on the assessor's $4,531,688. R0164287"],
]
fixed=[]
for i,r in enumerate(ies):
    row=at+i; fixed.append([r[0],r[1],r[2],f"=B{row}-C{row}",r[3]])
ins(ws,at,fixed,5,5,{2:MON,3:MON,4:MON})
wb.save(P); print("part 1 saved")
