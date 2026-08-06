# -*- coding: utf-8 -*-
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
from copy import copy
P="/home/user/Expense-Lease-Extractor-/Master_Property_Database.xlsx"
wb=load_workbook(P); MON='"$"#,##0.00'
def sf_(ws,r,nc): return [copy(ws.cell(row=r,column=c)._style) for c in range(1,nc+1)]
def put(ws,row,vals,st,nf=None):
    for j,v in enumerate(vals,1):
        c=ws.cell(row=row,column=j); c.value=v
        if j-1<len(st): c._style=copy(st[j-1])
        if nf and j in nf: c.number_format=nf[j]
        c.alignment=Alignment(vertical="top",wrap_text=True)
ws=wb["Income-Expense Detail"]; D=[]
def blk(p,c,items,printed=None,label="TOTAL",note=""): D.append((p,c,items,printed,label,note))

blk("5690 E. 56th Avenue - CY2023","Expense",[
 ("Real Estate Taxes",116160,"$3.32 PSF"),("CAM - Reimbursable",51894,"$1.48 PSF"),
 ("CAM - Non-Reimbursable",1085,"$0.03 PSF"),("Reserves for Replacement",0,"Reported as $0")],169139,"TOTAL EXPENSES")
blk("5690 E. 56th Avenue - CY2024","Expense",[
 ("Real Estate Taxes",123195,"$3.52 PSF"),("CAM - Reimbursable",65115,"$1.86 PSF"),
 ("CAM - Non-Reimbursable",6377,"$0.18 PSF"),("Reserves for Replacement",0,"Reported as $0")],194687,"TOTAL EXPENSES")
blk("Alpine Park II - CY2024","Income",[("Market Rent",229531.00,"No CAM reimbursements or misc income reported")],229531.00,"TOTAL INCOME")
blk("Alpine Park II - CY2024","Expense",[
 ("Real Property Taxes",69401.60,""),("Trash Services",8500.75,""),("Fire / Life & Safety R&M",8299.80,""),
 ("Exterior R&M",8000.00,""),("Water & Sewer",3438.85,""),("Landscape Services",3350.00,""),
 ("Property Insurance",1634.57,""),("Partnership Accounting",1550.00,""),("Snow Removal Services",1050.00,""),
 ("Maintenance & Repairs",819.88,""),("Electric",712.77,""),("Software",353.40,""),
 ("Merchant Account",221.71,""),("Interior Unit R&M",175.00,""),("Marketing Supplies",25.00,""),
 ("Partnership Admin",20.00,""),("Electrical R&M",0.00,""),("HVAC R&M",0.00,""),("Plumbing R&M",0.00,""),
 ("Bank Fees",-0.03,"Credit")],107553.30,"TOTAL OPERATING EXPENSES")
blk("Alpine Park II - CY2023","Income",[("Market Rent",256200.00,"")],256200.00,"TOTAL INCOME")
blk("Alpine Park II - CY2023","Expense",[
 ("Real Property Taxes",59985.48,""),("Trash Services",9561.90,""),("Property Insurance",5867.85,""),
 ("Fire / Life & Safety R&M",3870.75,""),("Snow Removal Services",2800.00,""),("Water & Sewer",2311.97,""),
 ("Landscape Services",2200.00,""),("Plumbing R&M",1094.49,""),("Electrical R&M",990.00,""),
 ("Partnership Accounting",800.00,""),("Electric",760.15,""),("Interior Unit R&M",618.00,""),
 ("HVAC R&M",480.00,""),("Software",316.05,""),("Merchant Account",281.54,"")],91938.18,"TOTAL OPERATING EXPENSES")
blk("Alpine Park II - CY2022","Income",[
 ("Market Rent",255992.67,""),("Other Income",2500.00,""),("Damages",500.00,""),("Interest Income",0.00,"")],258992.67,"TOTAL INCOME")
blk("Alpine Park II - CY2022","Expense",[
 ("Real Property Taxes",66104.78,""),("Trash Services",8303.97,""),("Property Insurance",5276.00,""),
 ("Landscape Services",4280.00,""),("Plumbing R&M",4115.00,""),("Water & Sewer",3043.11,""),
 ("Exterior R&M",2333.33,""),("Electric",887.54,""),("Interior Common Area R&M",780.00,""),
 ("Software",633.98,""),("Electrical R&M",611.00,""),("Bank Fees",25.00,""),
 ("Interior Unit R&M",0.00,""),("HVAC R&M",0.00,""),("Snow Removal Services",0.00,""),
 ("Merchant Account",0.00,"")],96393.71,"TOTAL OPERATING EXPENSES",
 "Printed total: $96,393.71. NOTE: the agent's own Income Analysis reports 2022 total income of $258,492 (omitting the $500 damages) and real estate taxes of $59,147 against the $66,104.78 on this statement.")
blk("Alpine Park II - CY2021","Income",[
 ("Market Rent",282764.52,""),("Other Income",3863.06,""),("Damages",2392.04,""),("Interest Income",116.90,"")],289136.52,"TOTAL INCOME")
blk("Alpine Park II - CY2021","Expense",[
 ("Real Property Taxes",48283.72,""),("Trash Services",6710.19,""),("Property Insurance",4893.00,""),
 ("Landscape Services",3875.00,""),("Water & Sewer",3344.39,""),("Interior Common Area R&M",2923.00,""),
 ("Plumbing R&M",1144.00,""),("Snow Removal Services",870.00,""),("Interior Unit R&M",846.00,""),
 ("Electric",648.33,""),("Merchant Account",408.51,""),("Software",322.98,""),
 ("HVAC R&M",280.00,""),("Exterior R&M",198.75,""),("Bank Fees",31.00,""),
 ("Electrical R&M",0.00,"")],74778.87,"TOTAL OPERATING EXPENSES")
for yr,base,rec,tot_i,op,tax,nonrec,tot_e in [
  ("YE2024",921507,1224772,2146279,394456,912402,1118,1307976),
  ("YE2023",921507,1227167,2148674,356482,923586,17221,1297289),
  ("YE2022",837490,797652,1635142,223270,533285,8751,765306)]:
    blk(f"Park 70 - {yr}","Income",[("Base Rent",base,""),("Recoverable Income",rec,""),
        ("Vacancy Loss",0,"100% occupied"),("Other Income",0,"")],tot_i,"EFFECTIVE GROSS INCOME")
    blk(f"Park 70 - {yr}","Expense",[("Property Taxes",tax,""),("Operating Expenses",op,""),
        ("Non-Recoverable Expenses",nonrec,""),("Management Fee",0,"Reported as $0"),
        ("Reserves for Replacement",0,"Reported as $0")],tot_e,"TOTAL EXPENSES")
blk("2850 Walden Street - 2025 (owner schedule)","Income",[
 ("Scheduled rent - single tenant",210589,"$7.57 PSF on 27,819 SF")],210589,"TOTAL INCOME")
blk("2850 Walden Street - 2025 (owner schedule)","Expense",[
 ("Real property tax (the only owner-borne expense)",158923,"Tenant pays all other expenses and improvements per the lease")],158923,"TOTAL EXPENSES")

at=ws.max_row+1; st=sf_(ws,5,5); CHECK=[]; row=at
for p,c,items,printed,label,note in D:
    s=row
    for n_,a,nt in items: put(ws,row,[p,c,n_,a,nt],st,{4:MON}); row+=1
    pn=(f"Printed total: ${printed:,.2f}" if printed is not None else "")
    put(ws,row,[p,c,label,f"=SUM(D{s}:D{row-1})",note or pn],st,{4:MON})
    ws.cell(row=row,column=3).font=Font(name="Arial",size=10,bold=True)
    ws.cell(row=row,column=4).font=Font(name="Arial",size=10,bold=True)
    CHECK.append((p,c,sum(a for _,a,_ in items),printed)); row+=1

ws=wb["Related Parcels"]; at=ws.max_row+1; st=sf_(ws,5,5)
ws.cell(row=at,column=1,value="KEW Realty Corporation - 'KEW East, West & Park' industrial/flex buildings, Commerce City (2025/2026 LOA):").font=Font(name="Arial",size=10,bold=True,italic=True)
for i,r in enumerate([["KEW Realty Corporation","R0187787","5690 E. 56th Avenue, Commerce City","Adams County","CO"],
                      ["KEW Realty Corporation","R0187789","5710 E. 56th Avenue, Commerce City","Adams County","CO"]]):
    put(ws,at+1+i,r,st)
wb.save(P); print("SAVED\n")
print("%-46s %-8s %14s %14s %s"%("BLOCK","CAT","PY SUM","PRINTED","MATCH")); bad=0
for p,c,s,pr in CHECK:
    if pr is None: print("%-46s %-8s %14.2f %14s n/a"%(p[:46],c,s,"(none)")); continue
    d=abs(s-pr); ok=d<0.005
    tag="OK" if ok else ("ANNOTATED $%.0f source rounding"%d if d<=1.001 else "*** MISMATCH ***")
    if not ok and d>1.001: bad+=1
    print("%-46s %-8s %14.2f %14.2f %s"%(p[:46],c,s,pr,tag))
print("\nUNEXPLAINED MISMATCHES:",bad)
